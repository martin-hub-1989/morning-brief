#!/usr/bin/env python3
"""
从合并种子文件 seed.xlsx 导入外汇原始序列到 SQLite。

数据来自两个工作表：
  - Fixing: 美元兑人民币中间价、即期汇率（日频，1981 年起）
  - Fwd Spread: CNH 远期、在岸掉期点、中美短端利率（日频）

仅导入原始数据（可从 Wind MCP 每日更新）。衍生序列（汇率拆解、套保成本、年化）
由 recompute_fx_derived.py 从原始数据计算，不在此导入。

用法:
  python3 scripts/import_fx_data.py                     # 导入全部（幂等）
  python3 scripts/import_fx_data.py --dry-run           # 干跑
  python3 scripts/import_fx_data.py --verbose           # 详细输出
"""

import argparse
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

# Windows GBK 编码兼容：强制 stdout/stderr 使用 UTF-8
if sys.platform == 'win32':
    for _s in (sys.stdout, sys.stderr):
        try:
            _s.reconfigure(encoding='utf-8')
        except Exception:
            pass

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DB = ROOT / "data" / "morning_brief.sqlite"
DEFAULT_XLSX = ROOT / "seed" / "seed.xlsx"

# ── Series definitions ─────────────────────────────────────────────────
#
# update_method 约定:
#   "wind_mcp"   — 原始数据，未来可从 Wind MCP 定期拉取更新
#   "cfets"      — 原始数据，来源中国外汇交易中心，可从 Wind 获取
#   "derived"    — 非原始数据，由 Excel/Python 公式基于原始数据计算，不从 MCP 获取
#
# (series_id, display_name, sheet_name, unit, source_name, update_method, excel_sheet, excel_col)
FX_SERIES = [
    # ── 原始数据 (Wind / CFETS) — 仅含可直接从 Wind MCP 更新的原始序列 ──
    # 衍生序列（汇率拆解、套保成本、年化）由 recompute_fx_derived.py 从原始数据计算
    #
    # Sheet "Fixing" (cols 2-3)
    ("fx:usdcny-fixing",     "USDCNY中间价",       "外汇", "fx",      "CFETS", "cfets",    "Fixing", 2),
    ("fx:usdcny-spot",       "USDCNY即期汇率",     "外汇", "fx",      "CFETS", "cfets",    "Fixing", 3),
    # Sheet "Fwd Spread" (cols 2-10) — CNH spot + forwards + CNY swap points
    ("fx:usdcnh-spot",       "USDCNH即期汇率",     "外汇", "fx",      "Wind",  "wind_mcp", "Fwd Spread", 2),
    ("fx:cnh-df-1m",         "USDCNH DF 1M",       "外汇", "fx",      "Wind",  "wind_mcp", "Fwd Spread", 3),
    ("fx:cnh-df-3m",         "USDCNH DF 3M",       "外汇", "fx",      "Wind",  "wind_mcp", "Fwd Spread", 4),
    ("fx:cnh-df-6m",         "USDCNH DF 6M",       "外汇", "fx",      "Wind",  "wind_mcp", "Fwd Spread", 5),
    ("fx:cnh-df-1y",         "USDCNH DF 1Y",       "外汇", "fx",      "Wind",  "wind_mcp", "Fwd Spread", 6),
    ("fx:cny-swap-1m",       "USDCNY掉期点 1M",    "外汇", "price",   "Wind",  "wind_mcp", "Fwd Spread", 7),
    ("fx:cny-swap-3m",       "USDCNY掉期点 3M",    "外汇", "price",   "Wind",  "wind_mcp", "Fwd Spread", 8),
    ("fx:cny-swap-6m",       "USDCNY掉期点 6M",    "外汇", "price",   "Wind",  "wind_mcp", "Fwd Spread", 9),
    ("fx:cny-swap-1y",       "USDCNY掉期点 1Y",    "外汇", "price",   "Wind",  "wind_mcp", "Fwd Spread", 10),
    # Sheet "Fwd Spread" (cols 15, 19) — 中美短端利率
    ("fx:cny-bond-1y",       "中债国债 1Y",        "外汇", "percent_point","Wind","wind_mcp","Fwd Spread", 15),
    ("fx:usd-bond-1y",       "美国国债 1Y",        "外汇", "percent_point","Wind","wind_mcp","Fwd Spread", 19),
]


def log(msg, level="INFO"):
    prefix = {"INFO": "  ", "WARN": "  ⚠", "ERROR": "  ✗", "OK": "  ✓"}
    print(f"{prefix.get(level, '  ')} {msg}", file=sys.stderr if level == "ERROR" else sys.stdout)


def parse_date(v):
    """Parse Excel datetime to date string YYYY-MM-DD."""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        return v[:10]
    return None


def import_fx_data(db_path, xlsx_path, dry_run=False, verbose=False):
    if not Path(xlsx_path).exists():
        log(f"Excel not found: {xlsx_path}", "ERROR")
        return 0

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    conn = sqlite3.connect(db_path)
    imported_at = datetime.now().isoformat(timespec="seconds")

    # Ensure series rows exist
    for sid, display_name, sheet_name, unit, source_name, update_method, _, _ in FX_SERIES:
        conn.execute(
            """INSERT OR IGNORE INTO series (
                   series_id, display_name, sheet_name, frequency, unit,
                   source_name, source_code, active, update_method, created_at, updated_at
               ) VALUES (?, ?, ?, 'D', ?, ?, ?, 1, ?, ?, ?)""",
            (sid, display_name, sheet_name, unit, source_name, sid, update_method, imported_at, imported_at),
        )

    total_inserted = 0

    for sid, display_name, sheet_name, unit, source_name, update_method, excel_sheet, excel_col in FX_SERIES:
        ws = wb[excel_sheet]
        if ws is None:
            log(f"Sheet '{excel_sheet}' not found — skipping {sid}", "WARN")
            continue

        # Find data rows: row 5 onwards, col 1 = date, excel_col = value
        points = []
        for r in range(5, ws.max_row + 1):
            date_val = ws.cell(row=r, column=1).value
            val = ws.cell(row=r, column=excel_col).value
            d = parse_date(date_val)
            if d is None:
                continue
            if val is None:
                continue
            # Convert zero to None (0 = missing data in Wind exports)
            try:
                fv = float(val)
            except (ValueError, TypeError):
                continue
            if fv == 0:
                continue
            points.append((d, fv))

        if not points:
            log(f"No data for {sid}", "WARN")
            continue

        # Get existing dates for this series to avoid re-inserting
        existing = set()
        for row in conn.execute(
            "SELECT date FROM observations WHERE series_id = ?", (sid,)
        ).fetchall():
            existing.add(row[0])

        new_pts = [(d, v) for d, v in points if d not in existing]

        if verbose:
            log(f"{display_name}: {len(points)} pts total, {len(new_pts)} new")

        if not dry_run:
            for d, v in new_pts:
                conn.execute(
                    """INSERT OR REPLACE INTO observations (series_id, date, value, as_of_date, imported_at)
                       VALUES (?, ?, ?, ?, ?)""",
                    (sid, d, v, d, imported_at),
                )
        total_inserted += len(new_pts)

    wb.close()

    if not dry_run and total_inserted > 0:
        conn.commit()
        log(f"Committed {total_inserted} new observations", "OK")
    elif dry_run:
        log(f"[DRY RUN] Would insert {total_inserted} observations", "WARN")
    else:
        log("No new observations to insert")

    conn.close()
    return total_inserted


def main():
    parser = argparse.ArgumentParser(description="Import FX data from 中间价与套保成本.xlsx")
    parser.add_argument("--db", default=str(DEFAULT_DB), help="Path to SQLite database")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Path to Excel source")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    if not Path(args.db).exists():
        log(f"Database not found: {args.db}", "ERROR")
        sys.exit(1)

    log("Martin Morning Brief — import_fx_data.py")
    if args.dry_run:
        log("Mode: DRY RUN (no writes)", "WARN")

    count = import_fx_data(args.db, args.xlsx, dry_run=args.dry_run, verbose=args.verbose)
    log(f"=== FX Import Summary ===")
    log(f"New obs:  {count}")


if __name__ == "__main__":
    main()
