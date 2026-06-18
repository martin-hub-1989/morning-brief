#!/usr/bin/env python3
"""
从 seed.xlsx 导入全部历史数据到 SQLite。

支持的 sheet:
  - 走势图 / PE TTM / PB LF / 股息率 — 通用格式，自动识别列名
  - Fixing / Fwd Spread           — 外汇原始数据，固定列映射
  - 美元指数                        — DXY 月频 + 归一化周期计算

用法:
  python3 scripts/import_seed.py --replace        # 重建数据库（全部 sheet）
  python3 scripts/import_seed.py --dry-run        # 干跑
  python3 scripts/import_seed.py --verbose        # 详细输出
  python3 scripts/import_seed.py --skip-fx        # 跳过外汇 sheet
  python3 scripts/import_seed.py --skip-super-cycle  # 跳过美元指数 sheet
"""

import argparse
import math
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. pip install openpyxl")
    sys.exit(1)

from lib import ROOT, DEFAULT_DB, log, open_db

DEFAULT_XLSX = ROOT / "seed" / "seed.xlsx"

# ── Phase 1: 通用 sheet（走势图 / PE TTM / PB LF / 股息率）────────────

SHEET_PREFIX = {
    "走势图": "trend",
    "PE TTM": "pe_ttm",
    "PB LF": "pb_lf",
    "股息率": "dividend_yield",
}

SHEET_FREQUENCY = {
    "走势图": "D",
    "PE TTM": "D",
    "PB LF": "D",
    "股息率": "D",
}

DISPLAY_ALIASES = {
    "881001.WI": "万得全A",
    "000001.SH": "上证指数",
    "000510.SH": "中证A500",
    "000300.SH": "沪深300",
    "000905.SH": "中证500",
    "000852.SH": "中证1000",
    "932000.CSI": "中证2000",
    "000922.CSI": "中证红利",
    "399006.SZ": "创业板指",
    "000688.SH": "科创50",
    "HSI.HI": "恒生指数",
    "HSTECH.HI": "恒生科技",
    "HSHDYI.HI": "恒生高股息率",
    "IXIC.GI": "纳斯达克指数",
    "SPX.GI": "标普500",
    "NH0100.NHF": "南华商品指数",
    "H00300.CSI": "300收益",
    "H00905.CSI": "500收益",
    "H00852.SH": "中证1000全收益",
    "399606.SZ": "创业板R",
    "000688CNY01.SH": "科创50(全)",
    "H00922.CSI": "中证红利全收益",
    "CI005917.WI": "金融(风格.中信)",
    "CI005918.WI": "周期(风格.中信)",
    "CI005919.WI": "消费(风格.中信)",
    "CI005920.WI": "成长(风格.中信)",
    "CI005921.WI": "稳定(风格.中信)",
    "HSIRH.HI": "恒生指数R",
    "HSTECHT.HI": "恒生科技R",
    "HSI52.HI": "恒生高股息率R",
    "SP500TR.SPI": "标普500全收益指数",
    "XCMP.GI": "纳斯达克总回报指数",
    "USDCNY.IB": "USDCNY",
    "USDJPY.IB": "USDJPY",
    "USDHKD.IB": "USDHKD",
    "GBPCNY.IB": "GBPCNY",
    "AUDCNY.IB": "AUDCNY",
    "USDX.FX": "DXY",
    "CNYX.IB": "CNYX",
    "Au9999.SGE": "AUXCNY",
    "SPTAUUSDOZ.IDC": "伦敦金现",
    "B.IPE": "ICE布油",
}


def clean_text(value):
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    return str(value).strip().replace("\t", "")


def is_date(value):
    if isinstance(value, pd.Timestamp):
        return True
    parsed = pd.to_datetime(value, errors="coerce")
    return not pd.isna(parsed)


def first_data_row(raw):
    for idx, value in raw.iloc[:, 0].items():
        if is_date(value):
            return idx
    raise ValueError("No date column found")


def infer_unit(sheet, display, source):
    text = f"{display} {source}"
    if sheet in ("PE TTM", "PB LF"):
        return "multiple"
    if sheet == "股息率":
        return "percent"
    if "收益率" in text or "债" in display:
        return "percent_point"
    if any(x in display for x in ["USDCNY", "USDJPY", "USDHKD", "GBPCNY", "AUDCNY"]):
        return "fx"
    if any(x in display for x in ["金", "油"]):
        return "price"
    return "index"


def display_name(sheet, col, raw, data_start):
    top = clean_text(raw.iat[0, col]) if raw.shape[0] > 0 else ""
    second = clean_text(raw.iat[1, col]) if raw.shape[0] > 1 else ""
    source = clean_text(raw.iat[2, col]) if raw.shape[0] > 2 and data_start >= 3 else ""

    if top and top not in {"error!", "Wind"}:
        return top
    if not top and source:
        alias = DISPLAY_ALIASES.get(source)
        if alias:
            return alias
    if second and second not in {"error!", "Wind"}:
        return second
    return DISPLAY_ALIASES.get(source, source or f"column_{col}")


def source_name(sheet, col, raw, data_start):
    if data_start >= 3 and raw.shape[0] > 2:
        return clean_text(raw.iat[2, col])
    return clean_text(raw.iat[data_start - 1, col])


def setup_db(conn):
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS series (
            series_id TEXT PRIMARY KEY,
            display_name TEXT NOT NULL,
            sheet_name TEXT NOT NULL,
            frequency TEXT NOT NULL,
            unit TEXT NOT NULL,
            source_name TEXT,
            source_code TEXT,
            active INTEGER NOT NULL DEFAULT 1,
            update_method TEXT NOT NULL DEFAULT 'manual',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS observations (
            series_id TEXT NOT NULL,
            date TEXT NOT NULL,
            value REAL NOT NULL,
            as_of_date TEXT,
            imported_at TEXT NOT NULL,
            PRIMARY KEY (series_id, date),
            FOREIGN KEY (series_id) REFERENCES series(series_id)
        );

        CREATE VIEW IF NOT EXISTS latest_observations AS
        SELECT o.series_id, s.display_name, s.sheet_name, s.frequency, s.unit,
               o.date, o.value
        FROM observations o
        JOIN series s ON s.series_id = o.series_id
        JOIN (
            SELECT series_id, MAX(date) AS date
            FROM observations
            GROUP BY series_id
        ) latest
          ON latest.series_id = o.series_id AND latest.date = o.date;
        """
    )


# ── Phase 1: import generic sheets ────────────────────────────────────

def _import_generic_sheets(conn, xlsx_path, imported_at, verbose=False):
    """导入走势图 / PE TTM / PB LF / 股息率 sheet。"""
    xls = pd.ExcelFile(xlsx_path)
    total_series = 0
    total_obs = 0

    for sheet in xls.sheet_names:
        if sheet not in SHEET_PREFIX:
            continue
        raw = pd.read_excel(xlsx_path, sheet_name=sheet, header=None)
        start = first_data_row(raw)
        dates = pd.to_datetime(raw.iloc[start:, 0], errors="coerce")
        prefix = SHEET_PREFIX[sheet]
        frequency = SHEET_FREQUENCY[sheet]

        for col in range(1, raw.shape[1]):
            values = pd.to_numeric(raw.iloc[start:, col], errors="coerce")
            if values.dropna().empty:
                continue

            label = display_name(sheet, col, raw, start)
            source = source_name(sheet, col, raw, start)
            series_id = f"{prefix}:{label}"
            now = imported_at
            unit = infer_unit(sheet, label, source)

            conn.execute(
                """INSERT INTO series (
                       series_id, display_name, sheet_name, frequency, unit,
                       source_name, source_code, created_at, updated_at
                   ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                   ON CONFLICT(series_id) DO UPDATE SET
                       display_name=excluded.display_name,
                       sheet_name=excluded.sheet_name,
                       frequency=excluded.frequency,
                       unit=excluded.unit,
                       source_name=excluded.source_name,
                       source_code=excluded.source_code,
                       updated_at=excluded.updated_at""",
                (series_id, label, sheet, frequency, unit, source, source, now, now),
            )
            total_series += 1

            for date_value, value in zip(dates, values):
                if pd.isna(date_value) or pd.isna(value):
                    continue
                conn.execute(
                    """INSERT INTO observations (series_id, date, value, as_of_date, imported_at)
                       VALUES (?, ?, ?, ?, ?)
                       ON CONFLICT(series_id, date) DO UPDATE SET
                           value=excluded.value,
                           as_of_date=excluded.as_of_date,
                           imported_at=excluded.imported_at""",
                    (series_id, date_value.date().isoformat(), float(value),
                     date_value.date().isoformat(), imported_at),
                )
                total_obs += 1

    return total_series, total_obs


# ── Phase 2: FX sheets (Fixing / Fwd Spread) ──────────────────────────

FX_SERIES = [
    ("fx:usdcny-fixing",     "USDCNY中间价",       "外汇", "fx",           "CFETS", "cfets",    "Fixing", 2),
    ("fx:usdcny-spot",       "USDCNY即期汇率",     "外汇", "fx",           "CFETS", "cfets",    "Fixing", 3),
    ("fx:usdcnh-spot",       "USDCNH即期汇率",     "外汇", "fx",           "Wind",  "wind_mcp", "Fwd Spread", 2),
    ("fx:cnh-df-1m",         "USDCNH DF 1M",       "外汇", "fx",           "Wind",  "wind_mcp", "Fwd Spread", 3),
    ("fx:cnh-df-3m",         "USDCNH DF 3M",       "外汇", "fx",           "Wind",  "wind_mcp", "Fwd Spread", 4),
    ("fx:cnh-df-6m",         "USDCNH DF 6M",       "外汇", "fx",           "Wind",  "wind_mcp", "Fwd Spread", 5),
    ("fx:cnh-df-1y",         "USDCNH DF 1Y",       "外汇", "fx",           "Wind",  "wind_mcp", "Fwd Spread", 6),
    ("fx:cny-swap-1m",       "USDCNY掉期点 1M",    "外汇", "price",        "Wind",  "wind_mcp", "Fwd Spread", 7),
    ("fx:cny-swap-3m",       "USDCNY掉期点 3M",    "外汇", "price",        "Wind",  "wind_mcp", "Fwd Spread", 8),
    ("fx:cny-swap-6m",       "USDCNY掉期点 6M",    "外汇", "price",        "Wind",  "wind_mcp", "Fwd Spread", 9),
    ("fx:cny-swap-1y",       "USDCNY掉期点 1Y",    "外汇", "price",        "Wind",  "wind_mcp", "Fwd Spread", 10),
    ("fx:cny-bond-1y",       "中债国债 1Y",        "外汇", "percent_point", "Wind",  "wind_mcp", "Fwd Spread", 15),
    ("fx:usd-bond-1y",       "美国国债 1Y",        "外汇", "percent_point", "Wind",  "wind_mcp", "Fwd Spread", 19),
]


def _parse_excel_date(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        return v[:10]
    return None


def _import_fx_sheets(conn, xlsx_path, imported_at, dry_run=False, verbose=False):
    """从 Fixing / Fwd Spread sheet 导入外汇原始数据。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    total = 0

    for sid, display_name, sheet_name, unit, source_name, update_method, excel_sheet, excel_col in FX_SERIES:
        ws = wb[excel_sheet]
        if ws is None:
            log(f"Sheet '{excel_sheet}' not found — skipping {sid}", "WARN")
            continue

        # Ensure series row exists
        conn.execute(
            """INSERT OR IGNORE INTO series (
                   series_id, display_name, sheet_name, frequency, unit,
                   source_name, source_code, active, update_method, created_at, updated_at
               ) VALUES (?, ?, ?, 'D', ?, ?, ?, 1, ?, ?, ?)""",
            (sid, display_name, sheet_name, unit, source_name, sid, update_method,
             imported_at, imported_at),
        )

        points = []
        for r in range(5, ws.max_row + 1):
            date_val = ws.cell(row=r, column=1).value
            val = ws.cell(row=r, column=excel_col).value
            d = _parse_excel_date(date_val)
            if d is None:
                continue
            if val is None:
                continue
            try:
                fv = float(val)
            except (ValueError, TypeError):
                continue
            if fv == 0:
                continue
            points.append((d, fv))

        if not points:
            log(f"No data for {sid}", "WARN")
            continue

        existing = set(
            row[0] for row in conn.execute(
                "SELECT date FROM observations WHERE series_id = ?", (sid,)
            ).fetchall()
        )
        new_pts = [(d, v) for d, v in points if d not in existing]

        if verbose:
            log(f"  {display_name}: {len(points)} pts total, {len(new_pts)} new")

        if not dry_run:
            for d, v in new_pts:
                conn.execute(
                    """INSERT OR REPLACE INTO observations (series_id, date, value, as_of_date, imported_at)
                       VALUES (?, ?, ?, ?, ?)""",
                    (sid, d, v, d, imported_at),
                )
        total += len(new_pts)

    wb.close()
    return total


# ── Phase 3: Super Cycle sheet (美元指数) ───────────────────────────────

CYCLE_BASE_DATES = {
    "1985": "1985-02-28",
    "2002": "2002-01-31",
    "2025": "2025-09-30",
}

SUPER_CYCLE_RAW = [
    ("super_cycle:dxy_monthly", "美元指数(月频)",       "美元超级周期", "index", 2),
    ("super_cycle:real_dxy",    "实际美元指数(月频)",   "美元超级周期", "index", 3),
    ("super_cycle:nominal_dxy", "名义美元指数(月频)",   "美元超级周期", "index", 4),
]

SUPER_CYCLE_DERIVED = [
    ("super_cycle:dxy_1985", "DXY归一化(1985周期)", "美元超级周期", "index"),
    ("super_cycle:dxy_2002", "DXY归一化(2002周期)", "美元超级周期", "index"),
    ("super_cycle:dxy_2025", "DXY归一化(2025周期)", "美元超级周期", "index"),
    ("super_cycle:dae_1985", "D/AE归一化(1985周期)", "美元超级周期", "index"),
    ("super_cycle:dae_2002", "D/AE归一化(2002周期)", "美元超级周期", "index"),
    ("super_cycle:dae_2025", "D/AE归一化(2025周期)", "美元超级周期", "index"),
]


def _import_super_cycle(conn, xlsx_path, imported_at, dry_run=False, verbose=False):
    """从 美元指数 sheet 导入原始月频数据 + 计算归一化周期序列。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if "美元指数" not in wb.sheetnames:
        wb.close()
        log("Sheet '美元指数' not found in seed.xlsx — skipping super cycle import", "WARN")
        return 0, 0

    ws = wb["美元指数"]

    # Ensure series rows
    for sid, display_name, sheet_name, unit, _ in SUPER_CYCLE_RAW:
        conn.execute(
            """INSERT OR IGNORE INTO series (
                   series_id, display_name, sheet_name, frequency, unit,
                   source_name, source_code, active, update_method, created_at, updated_at
               ) VALUES (?, ?, ?, 'M', ?, 'Wind', ?, 1, 'manual', ?, ?)""",
            (sid, display_name, sheet_name, unit, sid, imported_at, imported_at),
        )

    raw_total = 0
    for sid, display_name, sheet_name, unit, excel_col in SUPER_CYCLE_RAW:
        points = []
        for r in range(6, ws.max_row + 1):
            date_val = ws.cell(row=r, column=1).value
            val = ws.cell(row=r, column=excel_col).value
            d = _parse_excel_date(date_val)
            if d is None:
                continue
            if val is None:
                continue
            try:
                fv = float(val)
            except (ValueError, TypeError):
                continue
            if fv == 0:
                continue
            points.append((d, fv))

        if not points:
            log(f"No data for {sid}", "WARN")
            continue

        existing = set(
            row[0] for row in conn.execute(
                "SELECT date FROM observations WHERE series_id = ?", (sid,)
            ).fetchall()
        )
        new_pts = [(d, v) for d, v in points if d not in existing]

        if verbose:
            log(f"  {display_name}: {len(points)} pts total, {len(new_pts)} new")

        if not dry_run:
            for d, v in new_pts:
                conn.execute(
                    """INSERT OR REPLACE INTO observations (series_id, date, value, as_of_date, imported_at)
                       VALUES (?, ?, ?, ?, ?)""",
                    (sid, d, v, d, imported_at),
                )
        raw_total += len(new_pts)

    wb.close()

    # Compute normalized series
    derived_total = _compute_normalized(conn, imported_at, dry_run, verbose)

    return raw_total, derived_total


def _compute_normalized(conn, imported_at, dry_run=False, verbose=False):
    """从原始月度数据计算 6 个归一化超级周期序列。"""
    dxy_data = {}
    real_data = {}
    for row in conn.execute(
        "SELECT date, value FROM observations WHERE series_id = ? ORDER BY date",
        ("super_cycle:dxy_monthly",)
    ).fetchall():
        dxy_data[row[0]] = row[1]
    for row in conn.execute(
        "SELECT date, value FROM observations WHERE series_id = ? ORDER BY date",
        ("super_cycle:real_dxy",)
    ).fetchall():
        real_data[row[0]] = row[1]

    if not dxy_data:
        log("Missing super_cycle:dxy_monthly — cannot compute normalized series", "ERROR")
        return 0

    for sid, display_name, sheet_name, unit in SUPER_CYCLE_DERIVED:
        conn.execute(
            """INSERT OR IGNORE INTO series (
                   series_id, display_name, sheet_name, frequency, unit,
                   source_name, source_code, active, update_method, created_at, updated_at
               ) VALUES (?, ?, ?, 'M', ?, 'Python复算', ?, 1, 'derived', ?, ?)""",
            (sid, display_name, sheet_name, unit, sid, imported_at, imported_at),
        )

    total = 0
    for cycle_label, base_date in CYCLE_BASE_DATES.items():
        base_dxy = dxy_data.get(base_date)
        if base_dxy and base_dxy != 0:
            dxy_norm = {}
            for d, v in sorted(dxy_data.items()):
                if d >= base_date:
                    dxy_norm[d] = round(v / base_dxy * 100, 6)
            dxy_norm = dict(sorted(dxy_norm.items())[:18])

            sid = f"super_cycle:dxy_{cycle_label}"
            existing = set(
                row[0] for row in conn.execute(
                    "SELECT date FROM observations WHERE series_id = ?", (sid,)
                ).fetchall()
            )
            new_pts = [(d, v) for d, v in dxy_norm.items() if d not in existing]
            if new_pts and not dry_run:
                for d, v in new_pts:
                    conn.execute(
                        """INSERT OR REPLACE INTO observations (series_id, date, value, as_of_date, imported_at)
                           VALUES (?, ?, ?, ?, ?)""",
                        (sid, d, v, d, imported_at),
                    )
            total += len(new_pts)
            if verbose:
                log(f"  {sid}: {len(dxy_norm)} dates, {len(new_pts)} new")

        base_real = real_data.get(base_date)
        if base_real and base_real != 0:
            dae_norm = {}
            for d, v in sorted(real_data.items()):
                if d >= base_date:
                    dae_norm[d] = round(v / base_real * 100, 6)
            dae_norm = dict(sorted(dae_norm.items())[:18])

            sid = f"super_cycle:dae_{cycle_label}"
            existing = set(
                row[0] for row in conn.execute(
                    "SELECT date FROM observations WHERE series_id = ?", (sid,)
                ).fetchall()
            )
            new_pts = [(d, v) for d, v in dae_norm.items() if d not in existing]
            if new_pts and not dry_run:
                for d, v in new_pts:
                    conn.execute(
                        """INSERT OR REPLACE INTO observations (series_id, date, value, as_of_date, imported_at)
                           VALUES (?, ?, ?, ?, ?)""",
                        (sid, d, v, d, imported_at),
                    )
            total += len(new_pts)
            if verbose:
                log(f"  {sid}: {len(dae_norm)} dates, {len(new_pts)} new")

    return total


# ── Main ───────────────────────────────────────────────────────────────

def import_workbook(xlsx_path, db_path, replace=False, dry_run=False, verbose=False,
                    skip_fx=False, skip_super_cycle=False):
    """从 seed.xlsx 导入全部数据。"""
    db_path.parent.mkdir(parents=True, exist_ok=True)
    if replace and db_path.exists():
        db_path.unlink()

    imported_at = datetime.now().isoformat(timespec="seconds")

    with open_db(db_path) as conn:
        setup_db(conn)

        # Phase 1: Generic sheets (走势图 / PE TTM / PB LF / 股息率)
        log("Phase 1/3: Importing trend & valuation sheets...")
        series_count, obs_count = _import_generic_sheets(conn, xlsx_path, imported_at, verbose)
        if not dry_run:
            conn.commit()
        log(f"  {series_count} series, {obs_count} observations", "OK")

        # Phase 2: FX sheets
        if not skip_fx:
            log("Phase 2/3: Importing FX sheets (Fixing / Fwd Spread)...")
            fx_count = _import_fx_sheets(conn, xlsx_path, imported_at, dry_run, verbose)
            if not dry_run and fx_count > 0:
                conn.commit()
            log(f"  {fx_count} new observations", "OK" if fx_count > 0 else "WARN")
        else:
            log("Phase 2/3: FX sheets — skipped", "WARN")
            fx_count = 0

        # Phase 3: Super Cycle
        if not skip_super_cycle:
            log("Phase 3/3: Importing super cycle (美元指数)...")
            sc_raw, sc_derived = _import_super_cycle(conn, xlsx_path, imported_at, dry_run, verbose)
            if not dry_run and (sc_raw + sc_derived) > 0:
                conn.commit()
            log(f"  Raw: {sc_raw}, Normalized: {sc_derived}", "OK" if (sc_raw + sc_derived) > 0 else "WARN")
        else:
            log("Phase 3/3: Super cycle — skipped", "WARN")
            sc_raw, sc_derived = 0, 0

    total_obs = obs_count + fx_count + sc_raw + sc_derived
    total_series = series_count + len(FX_SERIES) + len(SUPER_CYCLE_RAW) + len(SUPER_CYCLE_DERIVED)

    if dry_run:
        log(f"[DRY RUN] Would import {total_series} series, {total_obs} observations", "WARN")
    else:
        log(f"Imported {total_series} series, {total_obs} observations into {db_path}", "OK")

    return total_series, total_obs


def main():
    parser = argparse.ArgumentParser(description="Import all seed data from seed.xlsx into SQLite.")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Path to Excel source")
    parser.add_argument("--db", default=str(DEFAULT_DB), help="Path to SQLite database")
    parser.add_argument("--replace", action="store_true",
                        help="Rebuild database from scratch")
    parser.add_argument("--dry-run", action="store_true", help="Dry run, no writes")
    parser.add_argument("--verbose", "-v", action="store_true", help="Verbose output")
    parser.add_argument("--skip-fx", action="store_true", help="Skip FX sheets")
    parser.add_argument("--skip-super-cycle", action="store_true",
                        help="Skip super cycle (美元指数) sheet")
    args = parser.parse_args()

    if not Path(args.xlsx).exists():
        log(f"Excel not found: {args.xlsx}", "ERROR")
        sys.exit(1)

    log("Martin Morning Brief — import_seed.py (all-in-one)")
    log(f"Excel: {args.xlsx}")
    if args.dry_run:
        log("Mode: DRY RUN (no writes)", "WARN")
    if args.replace:
        log("Mode: REPLACE (rebuild database)", "WARN")

    import_workbook(
        Path(args.xlsx), Path(args.db),
        replace=args.replace, dry_run=args.dry_run, verbose=args.verbose,
        skip_fx=args.skip_fx, skip_super_cycle=args.skip_super_cycle,
    )


if __name__ == "__main__":
    main()
