#!/usr/bin/env python3
"""
从 Super Dollar Scenario.xlsx 导入美元超级周期数据到 SQLite。

导入内容：
  A) 原始月度数据（3 个序列）：
     - super_cycle:dxy_monthly   — 美元指数月频
     - super_cycle:real_dxy      — 实际美元指数:对国外发达经济体（月频）
     - super_cycle:nominal_dxy   — 名义美元指数:对国外发达经济体（月频）

  B) 衍生归一化数据（6 个序列，从原始数据计算）：
     - super_cycle:dxy_1985/2002/2025  — DXY 归一化到峰值的 3 个周期
     - super_cycle:dae_1985/2002/2025  — D/AE 归一化到峰值的 3 个周期

用法:
  python3 scripts/import_super_cycle.py                           # 导入全部
  python3 scripts/import_super_cycle.py --dry-run                 # 干跑
  python3 scripts/import_super_cycle.py --xlsx ~/Downloads/...    # 指定 Excel 路径
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. pip install openpyxl")
    sys.exit(1)

from lib import ROOT, DEFAULT_DB, log, open_db

DEFAULT_XLSX = ROOT / "seed" / "Super Dollar Scenario.xlsx"

# 三个美元周期峰值日期
CYCLE_BASE_DATES = {
    "1985": "1985-02-28",
    "2002": "2002-01-31",
    "2025": "2025-09-30",
}

# 原始月度序列定义
RAW_SERIES = [
    ("super_cycle:dxy_monthly", "美元指数(月频)", "美元超级周期", "index", 2),
    ("super_cycle:real_dxy",    "实际美元指数(月频)", "美元超级周期", "index", 3),
    ("super_cycle:nominal_dxy", "名义美元指数(月频)", "美元超级周期", "index", 4),
]

# 衍生归一化序列定义
DERIVED_SERIES = [
    ("super_cycle:dxy_1985", "DXY归一化(1985周期)", "美元超级周期", "index"),
    ("super_cycle:dxy_2002", "DXY归一化(2002周期)", "美元超级周期", "index"),
    ("super_cycle:dxy_2025", "DXY归一化(2025周期)", "美元超级周期", "index"),
    ("super_cycle:dae_1985", "D/AE归一化(1985周期)", "美元超级周期", "index"),
    ("super_cycle:dae_2002", "D/AE归一化(2002周期)", "美元超级周期", "index"),
    ("super_cycle:dae_2025", "D/AE归一化(2025周期)", "美元超级周期", "index"),
]


def parse_date(v):
    """Parse Excel datetime to date string YYYY-MM-DD."""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        return v[:10]
    return None


def import_raw_data(conn, xlsx_path, imported_at, dry_run=False, verbose=False):
    """从 Excel 导入原始月度 DXY/Real/Nominal 数据。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["美元指数"]

    # Ensure series rows exist
    for sid, display_name, sheet_name, unit, _ in RAW_SERIES:
        conn.execute(
            """INSERT OR IGNORE INTO series (
                   series_id, display_name, sheet_name, frequency, unit,
                   source_name, source_code, active, update_method, created_at, updated_at
               ) VALUES (?, ?, ?, 'M', ?, 'Wind', ?, 1, 'manual', ?, ?)""",
            (sid, display_name, sheet_name, unit, sid, imported_at, imported_at),
        )

    total = 0
    for sid, display_name, sheet_name, unit, excel_col in RAW_SERIES:
        points = []
        for r in range(6, ws.max_row + 1):
            date_val = ws.cell(row=r, column=1).value
            val = ws.cell(row=r, column=excel_col).value
            d = parse_date(date_val)
            if d is None:
                continue
            if val is None:
                continue
            try:
                fv = float(val)
            except (ValueError, TypeError):
                continue
            if fv == 0:
                continue  # Real/Nominal DXY 在 2006 年前为 0
            points.append((d, fv))

        if not points:
            log(f"No data for {sid}", "WARN")
            continue

        # Filter to existing dates
        existing = set(
            row[0] for row in conn.execute(
                "SELECT date FROM observations WHERE series_id = ?", (sid,)
            ).fetchall()
        )
        new_pts = [(d, v) for d, v in points if d not in existing]

        if verbose:
            log(f"{display_name}: {len(points)} pts total, {len(new_pts)} new")

        if not dry_run:
            for d, v in new_pts:
                conn.execute(
                    """INSERT OR REPLACE INTO observations (series_id, date, value, as_of_date, imported_at)
                       VALUES (?, ?, ?, ?, ?)""",
                    (sid, d, v, d, imported_at),
                )
        total += len(new_pts)

    wb.close()
    return total


def compute_normalized(conn, imported_at, dry_run=False, verbose=False):
    """从原始月度数据计算 6 个归一化超级周期序列。

    公式: normalized[date] = raw[date] / raw[base_date] * 100
    每个周期取 base_date 起 18 个月数据。
    """
    # Load raw monthly data
    dxy_data = {}
    real_data = {}
    for row in conn.execute(
        "SELECT date, value FROM observations WHERE series_id = ? ORDER BY date",
        ("super_cycle:dxy_monthly",)
    ).fetchall():
        dxy_data[row[0]] = row[1]
    for row in conn.execute(
        "SELECT date, value FROM observations WHERE series_id = ? ORDER BY date",
        ("super_cycle:real_dxy",)
    ).fetchall():
        real_data[row[0]] = row[1]

    if not dxy_data:
        log("Missing super_cycle:dxy_monthly — cannot compute normalized series", "ERROR")
        return 0

    # Ensure derived series rows exist
    for sid, display_name, sheet_name, unit in DERIVED_SERIES:
        conn.execute(
            """INSERT OR IGNORE INTO series (
                   series_id, display_name, sheet_name, frequency, unit,
                   source_name, source_code, active, update_method, created_at, updated_at
               ) VALUES (?, ?, ?, 'M', ?, 'Python复算', ?, 1, 'derived', ?, ?)""",
            (sid, display_name, sheet_name, unit, sid, imported_at, imported_at),
        )

    total = 0
    for cycle_label, base_date in CYCLE_BASE_DATES.items():
        # DXY normalized
        base_dxy = dxy_data.get(base_date)
        if base_dxy and base_dxy != 0:
            dxy_norm = {}
            for d, v in sorted(dxy_data.items()):
                if d >= base_date:
                    dxy_norm[d] = round(v / base_dxy * 100, 6)
            # Take first 18 months
            dxy_norm = dict(sorted(dxy_norm.items())[:18])

            sid = f"super_cycle:dxy_{cycle_label}"
            count = _upsert_normalized(conn, sid, dxy_norm, imported_at, dry_run)
            total += count
            if verbose:
                log(f"  {sid}: {len(dxy_norm)} dates, {count} new")

        # D/AE normalized (uses Real DXY)
        base_real = real_data.get(base_date)
        if base_real and base_real != 0:
            dae_norm = {}
            for d, v in sorted(real_data.items()):
                if d >= base_date:
                    dae_norm[d] = round(v / base_real * 100, 6)
            dae_norm = dict(sorted(dae_norm.items())[:18])

            sid = f"super_cycle:dae_{cycle_label}"
            count = _upsert_normalized(conn, sid, dae_norm, imported_at, dry_run)
            total += count
            if verbose:
                log(f"  {sid}: {len(dae_norm)} dates, {count} new")

    return total


def _upsert_normalized(conn, series_id, data, imported_at, dry_run=False):
    """将归一化数据写入 observations（幂等）。"""
    existing = set(
        row[0] for row in conn.execute(
            "SELECT date FROM observations WHERE series_id = ?", (series_id,)
        ).fetchall()
    )
    new_pts = [(d, v) for d, v in data.items() if d not in existing]

    if new_pts and not dry_run:
        for d, v in new_pts:
            conn.execute(
                """INSERT OR REPLACE INTO observations (series_id, date, value, as_of_date, imported_at)
                   VALUES (?, ?, ?, ?, ?)""",
                (series_id, d, v, d, imported_at),
            )
    return len(new_pts)


def main():
    parser = argparse.ArgumentParser(description="Import Dollar Super Cycle data from Excel")
    parser.add_argument("--db", default=str(DEFAULT_DB), help="Path to SQLite database")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Path to Super Dollar Scenario.xlsx")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        # Try seed.xlsx (merged version)
        seed_path = ROOT / "seed" / "seed.xlsx"
        if seed_path.exists():
            log(f"Trying seed.xlsx instead of {args.xlsx}...")
            xlsx_path = seed_path
        else:
            log(f"Excel not found: {args.xlsx} (also tried {seed_path})", "ERROR")
            sys.exit(1)
    # Verify the "美元指数" sheet exists in whatever workbook we use
    wb_check = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "美元指数" not in wb_check.sheetnames:
        wb_check.close()
        log(f"'美元指数' sheet not found in {xlsx_path}", "ERROR")
        sys.exit(1)
    wb_check.close()

    if not Path(args.db).exists():
        log(f"Database not found: {args.db}", "ERROR")
        sys.exit(1)

    log("Martin Morning Brief — import_super_cycle.py")
    log(f"Excel: {xlsx_path}")
    if args.dry_run:
        log("Mode: DRY RUN (no writes)", "WARN")

    imported_at = datetime.now().isoformat(timespec="seconds")

    with open_db(args.db) as conn:
        # Phase 1: Import raw monthly data
        raw_count = import_raw_data(conn, xlsx_path, imported_at,
                                    dry_run=args.dry_run, verbose=args.verbose)
        if not args.dry_run and raw_count > 0:
            conn.commit()
        log(f"Raw data: {raw_count} new observations", "OK" if raw_count > 0 else "WARN")

        # Phase 2: Compute normalized series
        derived_count = compute_normalized(conn, imported_at,
                                           dry_run=args.dry_run, verbose=args.verbose)
        if not args.dry_run and derived_count > 0:
            conn.commit()
        log(f"Normalized: {derived_count} new observations", "OK" if derived_count > 0 else "WARN")

    log("=== Super Cycle Import Summary ===")
    log(f"Raw obs:      {raw_count}")
    log(f"Derived obs:  {derived_count}")


if __name__ == "__main__":
    main()
