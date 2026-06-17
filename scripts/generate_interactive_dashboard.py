#!/usr/bin/env python3
import argparse
import json
import math
import sqlite3
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DB = ROOT / "data" / "morning_brief.sqlite"
DEFAULT_OUTPUT = ROOT / "output" / "interactive_dashboard.html"

TREND_SERIES = [
    "trend:美-10债", "trend:中-10债", "trend:日-10债", "trend:英-10债", "trend:德-10债",
    "trend:法-10债", "trend:意-10债", "trend:澳-10债", "trend:中-30债", "trend:日-30债",
    "trend:美-30债", "trend:中-2债", "trend:美-2债",
    "trend:USDCNY", "trend:USDJPY", "trend:USDHKD", "trend:GBPCNY", "trend:AUDCNY",
    "trend:DXY", "trend:CNYX",
    "trend:万得全A", "trend:上证指数", "trend:中证A500", "trend:沪深300", "trend:中证500",
    "trend:中证1000", "trend:中证2000", "trend:中证红利", "trend:创业板指", "trend:科创50",
    "trend:恒生高股息率", "trend:恒生指数", "trend:恒生科技", "trend:纳斯达克指数", "trend:标普500",
    "trend:AUXCNY", "trend:伦敦金现", "trend:ICE布油", "trend:南华商品指数",
]

RETURN_SERIES = [
    {"series_id": "trend:300收益", "region": "A股"},
    {"series_id": "trend:500收益", "region": "A股"},
    {"series_id": "trend:中证1000全收益", "region": "A股"},
    {"series_id": "trend:创业板R", "region": "A股"},
    {"series_id": "trend:科创50(全)", "region": "A股"},
    {"series_id": "trend:中证红利全收益", "region": "A股"},
    {"series_id": "trend:恒生指数R", "region": "港股"},
    {"series_id": "trend:恒生科技R", "region": "港股"},
    {"series_id": "trend:恒生高股息率R", "region": "港股"},
    {"series_id": "trend:标普500全收益指数", "region": "美股"},
    {"series_id": "trend:纳斯达克总回报指数", "region": "美股"},
    {"series_id": "trend:AUXCNY", "region": "商品"},
    {"series_id": "trend:伦敦金现", "region": "商品"},
    {"series_id": "trend:ICE布油", "region": "商品"},
    {"series_id": "trend:南华商品指数", "region": "商品"},
    {"series_id": "trend:金融(风格.中信)", "region": "A股风格"},
    {"series_id": "trend:周期(风格.中信)", "region": "A股风格"},
    {"series_id": "trend:消费(风格.中信)", "region": "A股风格"},
    {"series_id": "trend:成长(风格.中信)", "region": "A股风格"},
    {"series_id": "trend:稳定(风格.中信)", "region": "A股风格"},
    {"series_id": "trend:万得全A", "region": "A股风格"},
]

EMOTION_SERIES = [
    "htsc:A股情绪指数",
    "htsc:港股情绪指数",
]

CAPITAL_SERIES = [
    "htsc:ETF资金",
    "htsc:融资资金",
    "htsc:公募基金",
    "htsc:散户资金净流入",
    "htsc:产业资本减持",
    "htsc:一级市场",
]

RATE_SERIES = [
    "trend:美-10债", "trend:中-10债", "trend:日-10债", "trend:英-10债", "trend:德-10债",
    "trend:法-10债", "trend:意-10债", "trend:澳-10债", "trend:中-30债", "trend:日-30债",
    "trend:美-30债", "trend:中-2债", "trend:美-2债",
]

FX_SERIES = [
    {"series_id": "trend:USDCNY", "group": "兑人民币"},
    {"series_id": "trend:GBPCNY", "group": "兑人民币"},
    {"series_id": "trend:AUDCNY", "group": "兑人民币"},
    {"series_id": "trend:USDJPY", "group": "美元交叉"},
    {"series_id": "trend:USDHKD", "group": "美元交叉"},
    {"series_id": "trend:DXY", "group": "指数"},
    {"series_id": "trend:CNYX", "group": "指数"},
]

SPREAD_SERIES = [
    "trend:中-10债", "trend:美-10债", "trend:中-2债", "trend:美-2债", "trend:USDCNY"
]

FX_FIXING_SERIES = [
    "fx:usdcny-fixing",
    "fx:usdcny-spot",
    "fx:decomp-night-20d",
    "fx:decomp-day-20d",
]

FX_COST_SERIES = [
    "fx:cny-hedge-1m", "fx:cny-hedge-3m", "fx:cny-hedge-6m", "fx:cny-hedge-1y",
    "fx:cnh-hedge-1m", "fx:cnh-hedge-3m", "fx:cnh-hedge-6m", "fx:cnh-hedge-1y",
    "fx:cny-hedge-ann-1m", "fx:cny-hedge-ann-3m", "fx:cny-hedge-ann-6m", "fx:cny-hedge-ann-1y",
    "fx:cnh-hedge-ann-1m", "fx:cnh-hedge-ann-3m", "fx:cnh-hedge-ann-6m", "fx:cnh-hedge-ann-1y",
    "fx:cny-bond-1y", "fx:usd-bond-1y",
]

VALUATION_PREFIXES = {
    "PE TTM": "pe_ttm",
    "PB LF": "pb_lf",
    "股息率": "dividend_yield",
}


def fill_zero_with_previous(points):
    filled = []
    previous = None
    for d, value in points:
        if value is None or math.isnan(value):
            continue
        if value == 0:
            if previous is None:
                continue
            value = previous
        else:
            previous = value
        filled.append([d, value])
    return filled


def load_metadata(conn, series_ids):
    if not series_ids:
        return {}
    q = ",".join("?" for _ in series_ids)
    rows = conn.execute(
        f"""
        SELECT series_id, display_name, sheet_name, frequency, unit, source_code
        FROM series
        WHERE series_id IN ({q})
        """,
        series_ids,
    ).fetchall()
    return {
        sid: {
            "display_name": name,
            "sheet_name": sheet,
            "frequency": freq,
            "unit": unit,
            "source_code": code,
        }
        for sid, name, sheet, freq, unit, code in rows
    }


def load_observations(conn, series_ids):
    if not series_ids:
        return {}
    q = ",".join("?" for _ in series_ids)
    rows = conn.execute(
        f"""
        SELECT series_id, date, value
        FROM observations
        WHERE series_id IN ({q})
        ORDER BY series_id, date
        """,
        series_ids,
    ).fetchall()
    grouped = {}
    for sid, d, value in rows:
        grouped.setdefault(sid, []).append((d, float(value)))
    # 零值不再预填，保留原始数据。
    # 单一指标图：跳过零值（视为当日无数据，避免影响后续均线）。
    # 双指标图：在 JS 层按需前向填补，保证两条序列时间轴对齐。
    return grouped


def load_valuation_options(conn):
    rows = conn.execute(
        """
        SELECT series_id, display_name, sheet_name
        FROM series
        WHERE sheet_name IN ('PE TTM', 'PB LF', '股息率')
          AND active = 1
          AND display_name != '万得全A指数:收盘价'
        ORDER BY display_name, sheet_name
        """
    ).fetchall()
    options = {}
    series_ids = []
    for sid, name, sheet in rows:
        options.setdefault(name, {})[sheet] = sid
        series_ids.append(sid)
    complete = [
        {"display_name": name, "metrics": metrics}
        for name, metrics in options.items()
        if set(metrics.keys()) == set(VALUATION_PREFIXES.keys())
    ]
    complete.sort(key=lambda item: item["display_name"])
    return complete, series_ids


def extract_super_cycle_data(xlsx_path=None):
    """Extract Dollar Super Cycle normalized data from Excel.

    Returns dict with {t_labels, dxy: {1985,2002,2025}, dae: {1985,2002,2025}}
    or None if extraction fails.
    """
    if not HAS_OPENPYXL:
        return None
    if xlsx_path is None:
        xlsx_path = Path.home() / "Downloads" / "Super Dollar Scenario.xlsx"
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        return None
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb["美元指数"]
        t_labels = []
        dxy = {"1985": [], "2002": [], "2025": []}
        dae = {"1985": [], "2002": [], "2025": []}
        for r in range(5, 40):
            t = ws.cell(row=r, column=10).value
            if t is None or not str(t).startswith("T+"):
                break
            t_labels.append(str(t))
            for col, target in [(14, dxy["1985"]), (15, dxy["2002"]), (16, dxy["2025"]),
                                 (23, dae["1985"]), (24, dae["2002"]), (25, dae["2025"])]:
                v = ws.cell(row=r, column=col).value
                target.append(round(v, 2) if v is not None else None)
        wb.close()
        return {"t_labels": t_labels, "dxy": dxy, "dae": dae}
    except Exception:
        return None


def build_payload(db_path):
    conn = sqlite3.connect(db_path)
    valuation_options, valuation_series = load_valuation_options(conn)
    return_items = [item["series_id"] for item in RETURN_SERIES]
    fx_items = [item["series_id"] for item in FX_SERIES]
    all_series = sorted(set(TREND_SERIES + return_items + RATE_SERIES + fx_items + SPREAD_SERIES + valuation_series + EMOTION_SERIES + CAPITAL_SERIES + FX_FIXING_SERIES + FX_COST_SERIES))
    meta = load_metadata(conn, all_series)
    observations = load_observations(conn, all_series)
    conn.close()

    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "source": {
            "database": str(db_path),
            "zero_rule": "0 values are treated as missing and forward-filled from the previous valid observation for charts and calculations.",
        },
        "meta": meta,
        "observations": observations,
        "trend_options": [
            {"series_id": sid, **meta[sid]} for sid in TREND_SERIES if sid in meta
        ],
        "return_options": [
            {"series_id": item["series_id"], "region": item["region"], **meta[item["series_id"]]}
            for item in RETURN_SERIES
            if item["series_id"] in meta
        ],
        "rate_options": [
            {"series_id": sid, **meta[sid]} for sid in RATE_SERIES if sid in meta
        ],
        "fx_options": [
            {"series_id": item["series_id"], "group": item["group"], **meta[item["series_id"]]}
            for item in FX_SERIES
            if item["series_id"] in meta
        ],
        "valuation_options": valuation_options,
        "valuation_metrics": list(VALUATION_PREFIXES.keys()),
        "fx_fixing_options": [
            {"series_id": sid, **meta[sid]} for sid in FX_FIXING_SERIES if sid in meta
        ],
        "fx_cost_options": [
            {"series_id": sid, **meta[sid]} for sid in FX_COST_SERIES if sid in meta
        ],
        "super_cycle": extract_super_cycle_data(),
    }


HTML_TEMPLATE = r"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Daily Morning Brief - Dashboard</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:ital,wght@0,400;0,500;0,600;0,700;1,400&family=Noto+Sans+SC:wght@400;500;600;700&display=swap" rel="stylesheet">
  <style>
    :root {
      --bg: #F9F6F0;
      --panel: #FFFFFF;
      --ink: #2C2416;
      --muted: #8B8581;
      --line: #E8E2D8;
      --soft: #F5F1E9;
      --blue: #5B7F8A;
      --red: #C44B3B;
      --green: #6B8B7A;
      --gold: #B8974A;
      --purple: #8B6B7E;
      --radius-sm: 6px;
      --radius-md: 10px;
      --radius-lg: 14px;
      --shadow-sm: 0 1px 3px rgba(44,36,22,0.05);
      --shadow-md: 0 4px 16px rgba(44,36,22,0.07);
      --shadow-lg: 0 12px 32px rgba(44,36,22,0.10);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      background: var(--bg);
      color: var(--ink);
      font-family: "IBM Plex Sans", "Noto Sans SC", "PingFang SC", "Microsoft YaHei", sans-serif;
      font-weight: 400;
      letter-spacing: -0.01em;
      -webkit-font-smoothing: antialiased;
      -moz-osx-font-smoothing: grayscale;
    }
    header.top {
      background: var(--panel);
      border-bottom: 1px solid var(--line);
      padding: 16px 28px 12px;
      position: sticky;
      top: 0;
      z-index: 5;
      backdrop-filter: blur(8px);
      -webkit-backdrop-filter: blur(8px);
      background: rgba(255,255,255,0.90);
    }
    h1 { margin: 0; font-size: 22px; font-weight: 600; line-height: 1.2; letter-spacing: -0.02em; }
    .freshness { margin-top: 5px; color: var(--muted); font-size: 12px; font-weight: 500; }

    /* ── Cover ─────────────────────────────────── */
    .cover {
      min-height: calc(100vh - 120px);
      display: flex;
      align-items: center;
      justify-content: center;
      background: var(--bg);
      border-radius: var(--radius-lg);
    }
    .cover-content {
      text-align: center;
      padding: 56px 24px;
      max-width: 1100px;
    }
    .cover-eyebrow {
      font-size: 12px;
      font-weight: 500;
      color: var(--muted);
      letter-spacing: 3px;
      text-transform: uppercase;
      margin: 0 0 28px;
    }
    .cover-title {
      font-size: clamp(48px, 8vw, 80px);
      font-weight: 700;
      letter-spacing: -0.03em;
      color: var(--ink);
      margin: 0 0 20px;
      line-height: 1.05;
    }
    .cover-accent {
      width: 40px;
      height: 2px;
      background: var(--gold);
      margin: 0 auto 28px;
    }
    .gold-letter { color: var(--gold); }
    .cover-subtitle {
      font-size: 15px;
      font-weight: 400;
      color: var(--muted);
      margin: 0 0 8px;
    }
    .cover-date {
      font-size: 13px;
      color: var(--muted);
      opacity: 0.6;
      margin: 0 0 56px;
      font-weight: 400;
    }
    .cover-cards {
      display: flex;
      flex-wrap: wrap;
      justify-content: center;
      gap: 12px;
      max-width: 1000px;
      margin: 0 auto;
    }
    .cover-card {
      flex: 0 0 calc(25% - 9px);
      min-width: 172px;
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: var(--radius-md);
      padding: 18px 16px 16px;
      cursor: pointer;
      transition: border-color 0.15s ease, box-shadow 0.15s ease, background 0.15s ease;
      text-align: center;
      box-shadow: var(--shadow-sm);
    }
    .cover-card:hover {
      border-color: #C44B3B;
      box-shadow: var(--shadow-md);
      background: #FDFCF9;
    }
    .cover-card:active {
      background: var(--soft);
    }
    .cover-card:focus-visible {
      outline: 2px solid var(--blue);
      outline-offset: 2px;
      border-radius: var(--radius-md);
    }
    .cover-card-title {
      display: block;
      font-size: 15px;
      font-weight: 600;
      color: var(--ink);
      margin-bottom: 4px;
      letter-spacing: -0.01em;
    }
    .cover-card-desc {
      display: block;
      font-size: 11px;
      font-weight: 400;
      color: var(--muted);
    }
    @media (max-width: 768px) {
      .cover-card { flex: 0 0 calc(50% - 5px); min-width: 140px; }
      .cover-cards { gap: 10px; }
      .cover-title { font-size: clamp(36px, 10vw, 52px); }
      .cover-eyebrow { letter-spacing: 2px; }
      .cover-content { padding: 40px 16px; }
    }
    @media (max-width: 480px) {
      .cover-card { flex: 0 0 100%; min-width: 0; }
      .cover-cards { max-width: 300px; }
      .cover-title { font-size: clamp(30px, 11vw, 42px); }
      .cover-content { padding: 32px 16px; }
    }
    @media (prefers-reduced-motion: reduce) {
      .cover-card { transition: none; }
      .tab,
      .segmented button,
      select,
      input[type="date"] { transition: none; }
    }

    main {
      max-width: 1536px;
      margin: 0 auto;
      padding: 20px 28px 48px;
    }
    .tabs {
      display: flex;
      gap: 6px;
      flex-wrap: wrap;
      margin-bottom: 18px;
    }
    .tab {
      border: 1px solid var(--line);
      background: var(--panel);
      border-radius: var(--radius-sm);
      padding: 8px 14px;
      cursor: pointer;
      color: var(--muted);
      font-size: 13px;
      font-weight: 500;
      font-family: inherit;
      transition: background 0.15s ease, color 0.15s ease, border-color 0.15s ease, box-shadow 0.15s ease;
    }
    .tab:hover {
      background: var(--soft);
      color: var(--ink);
      border-color: var(--red);
    }
    .tab.active {
      background: var(--ink);
      color: #fff;
      border-color: var(--ink);
      box-shadow: var(--shadow-sm);
    }
    .tab:focus-visible {
      outline: 2px solid var(--blue);
      outline-offset: 2px;
    }
    .sub-tabs {
      margin-top: -4px;
      margin-bottom: 18px;
      padding: 8px 12px;
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: var(--radius-sm);
      display: none;
    }
    .sub-tabs.visible { display: flex; }
    .sub-tabs .tab {
      font-size: 12px;
      padding: 6px 12px;
      border: none;
      background: transparent;
      color: var(--muted);
      border-radius: 4px;
    }
    .sub-tabs .tab:hover {
      background: var(--soft);
      color: var(--ink);
    }
    .sub-tabs .tab.active {
      background: var(--blue);
      color: #fff;
      box-shadow: none;
    }
    .view { display: none; }
    .view.active { display: block; }
    .panel {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: var(--radius-md);
      margin-bottom: 16px;
      overflow: hidden;
      box-shadow: var(--shadow-sm);
    }
    .panel-head {
      padding: 18px 20px 12px;
      border-bottom: 1px solid var(--line);
    }
    h2 { margin: 0; font-size: 17px; font-weight: 600; line-height: 1.25; letter-spacing: -0.01em; padding-left: 12px; border-left: 2px solid var(--red); }
    h3 { margin: 14px 0 6px 20px; font-size: 14px; font-weight: 600; letter-spacing: -0.01em; }
    .panel-head p { margin: 6px 0 0; color: var(--muted); font-size: 12px; line-height: 1.5; }
    .controls {
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 14px;
      padding: 16px 20px;
      border-bottom: 1px solid var(--line);
      background: var(--soft);
    }
    .control-wide { grid-column: span 2; }
    label { display: block; color: var(--muted); font-size: 11px; font-weight: 500; margin-bottom: 6px; text-transform: uppercase; letter-spacing: 0.03em; }
    select, input[type="date"] {
      width: 100%;
      height: 36px;
      border: 1px solid var(--line);
      border-radius: var(--radius-sm);
      background: var(--panel);
      color: var(--ink);
      padding: 0 10px;
      font-size: 13px;
      font-family: inherit;
      transition: border-color 0.15s ease, box-shadow 0.15s ease;
    }
    select:focus, input[type="date"]:focus {
      outline: none;
      border-color: var(--blue);
      box-shadow: 0 0 0 3px rgba(91,127,138,0.15);
    }
    .segmented {
      display: flex;
      flex-wrap: wrap;
      gap: 5px;
    }
    .segmented button {
      height: 32px;
      border: 1px solid var(--line);
      background: var(--panel);
      color: var(--muted);
      border-radius: var(--radius-sm);
      padding: 0 12px;
      cursor: pointer;
      font-size: 12px;
      font-weight: 500;
      font-family: inherit;
      transition: background 0.15s ease, color 0.15s ease, border-color 0.15s ease;
    }
    .segmented button:hover {
      background: var(--soft);
      color: var(--ink);
    }
    .segmented button.active {
      background: var(--blue);
      color: #fff;
      border-color: var(--blue);
    }
    .segmented button:focus-visible {
      outline: 2px solid var(--blue);
      outline-offset: 2px;
    }
    .mode-checks {
      display: flex;
      flex-wrap: wrap;
      gap: 6px 14px;
    }
    .mode-check {
      display: inline-flex;
      align-items: center;
      gap: 4px;
      font-size: 13px;
      color: #5B5550;
      cursor: pointer;
      user-select: none;
      padding: 4px 0;
    }
    .mode-check input[type="checkbox"] {
      accent-color: var(--blue);
      width: 15px;
      height: 15px;
      margin: 0;
      cursor: pointer;
    }
    .mode-check span {
      white-space: nowrap;
    }
    .data-label {
      font-size: 9px;
      font-family: "IBM Plex Sans", "Noto Sans SC", sans-serif;
      font-weight: 500;
    }
    .chart-area {
      padding: 14px 18px 6px;
      overflow-x: auto;
    }
    svg.chart {
      width: 100%;
      min-width: 900px;
      height: 420px;
      display: block;
    }
    .grid-line { stroke: #EAE4DA; stroke-width: 1; }
    .axis-line { stroke: #D4CCC0; stroke-width: 1; }
    .axis-text { fill: #8B8581; font-size: 11px; font-family: "IBM Plex Sans", "Noto Sans SC", sans-serif; }
    .legend-text { fill: #2C2416; font-size: 12px; font-weight: 500; }
    .chart-hit {
      fill: transparent;
      stroke: transparent;
      cursor: crosshair;
    }
    .chart-hit:hover {
      stroke: rgba(91,127,138,0.15);
      stroke-width: 2;
    }
    .tooltip-box {
      fill: #FFFFFF;
      stroke: #D4CCC0;
      stroke-width: 1;
      filter: drop-shadow(0 6px 14px rgba(44,36,22,0.10));
    }
    .tooltip-title {
      fill: #2C2416;
      font-size: 11px;
      font-weight: 600;
    }
    .tooltip-text {
      fill: #6B6560;
      font-size: 11px;
    }
    .table-wrap { overflow-x: auto; }
    table {
      width: 100%;
      border-collapse: collapse;
      font-size: 12px;
      font-variant-numeric: tabular-nums;
    }
    th, td {
      border-top: 1px solid var(--line);
      padding: 10px 12px;
      text-align: right;
      white-space: nowrap;
    }
    th:first-child, td:first-child { text-align: left; }
    th { color: var(--muted); background: var(--soft); font-weight: 600; font-size: 11px; text-transform: uppercase; letter-spacing: 0.03em; }
    .note {
      color: var(--muted);
      font-size: 11px;
      padding: 10px 20px 16px;
      line-height: 1.5;
    }
    .empty {
      padding: 80px 16px;
      text-align: center;
      color: var(--muted);
      font-size: 13px;
    }
    .sources {
      color: var(--muted);
      font-size: 11px;
      line-height: 1.6;
      padding: 16px 20px;
    }
    @media (max-width: 860px) {
      header.top { position: static; padding: 14px 16px; }
      main { padding: 12px 14px 32px; }
      .controls { grid-template-columns: 1fr; }
      .control-wide { grid-column: span 1; }
      h1 { font-size: 20px; }
      .tab { flex: 1; font-size: 12px; padding: 8px 10px; }
      .cover-content { padding: 36px 16px; }
    }

    /* ── Topics ─────────────────────────────────── */
    .topic-chart-area {
      padding: 14px 18px 6px;
      overflow-x: auto;
    }
    .topic-chart-area svg {
      width: 100%;
      min-width: 840px;
      height: 400px;
      display: block;
    }
    .topic-legend {
      display: flex;
      flex-wrap: wrap;
      gap: 16px;
      padding: 0 20px 12px;
      font-size: 12px;
      color: var(--muted);
    }
    .topic-legend span {
      display: flex;
      align-items: center;
      gap: 6px;
    }
    .topic-legend-swatch {
      display: inline-block;
      width: 20px;
      height: 2.5px;
      border-radius: 2px;
      flex-shrink: 0;
    }
  </style>
</head>
<body>
  <header class="top">
    <h1>Daily Morning Brief - Dashboard</h1>
    <div class="freshness" id="freshness"></div>
  </header>
  <main>
    <nav class="tabs primary-tabs" id="primary-tabs" aria-label="Primary sections">
      <button class="tab" data-group="trend">走势看板</button>
      <button class="tab" data-group="review">涨跌复盘</button>
      <button class="tab" data-group="equity">权益看板</button>
      <button class="tab" data-group="fxwatch">外汇看板</button>
    </nav>
    <nav class="tabs sub-tabs" id="sub-tabs" aria-label="Sub sections"></nav>

    <section class="view active" id="view-cover">
      <div class="cover">
        <div class="cover-content">
          <p class="cover-eyebrow">Designed by MARTIN</p>
          <h1 class="cover-title"><span class="gold-letter">M</span>ulti-<span class="gold-letter">A</span>sset Morning Brief</h1>
          <div class="cover-accent"></div>
          <p class="cover-subtitle">Daily Dashboard</p>
          <p class="cover-date" id="cover-date"></p>
          <div class="cover-cards">
            <div class="cover-card" data-view="trend">
              <span class="cover-card-title">走势看板</span>
              <span class="cover-card-desc">Multi-Asset Trends</span>
            </div>
            <div class="cover-card" data-view="returns">
              <span class="cover-card-title">股票涨跌</span>
              <span class="cover-card-desc">Equity Performance</span>
            </div>
            <div class="cover-card" data-view="rates">
              <span class="cover-card-title">利率涨跌</span>
              <span class="cover-card-desc">Rates Monitor</span>
            </div>
            <div class="cover-card" data-view="fx">
              <span class="cover-card-title">汇率涨跌</span>
              <span class="cover-card-desc">FX Movements</span>
            </div>
            <div class="cover-card" data-view="spread">
              <span class="cover-card-title">中美利差</span>
              <span class="cover-card-desc">CN-US Spreads</span>
            </div>
            <div class="cover-card" data-view="valuation">
              <span class="cover-card-title">估值看板</span>
              <span class="cover-card-desc">Valuation Lens</span>
            </div>
            <div class="cover-card" data-view="emotion">
              <span class="cover-card-title">市场情绪</span>
              <span class="cover-card-desc">Market Sentiment</span>
            </div>
            <div class="cover-card" data-view="fixing">
              <span class="cover-card-title">中间价</span>
              <span class="cover-card-desc">Central Parity</span>
            </div>
            <div class="cover-card" data-view="cost">
              <span class="cover-card-title">套保成本</span>
              <span class="cover-card-desc">Hedge Cost</span>
            </div>
            <div class="cover-card" data-view="topics">
              <span class="cover-card-title">专题图表</span>
              <span class="cover-card-desc">Topics</span>
            </div>
          </div>
        </div>
      </div>
    </section>

    <section class="view" id="view-trend">
      <div class="panel">
        <div class="panel-head">
          <h2>走势看板</h2>
          <p>选择 1 个或 2 个指标；选择 2 个时自动左右分轴。Y 轴按所选区间最高点和最低点落在 85% / 15% 位置自动缩放。</p>
        </div>
        <div class="controls">
          <div>
            <label for="trend-one">指标一</label>
            <select id="trend-one"></select>
          </div>
          <div>
            <label for="trend-two">指标二</label>
            <select id="trend-two"></select>
          </div>
          <div class="control-wide">
            <label>展示周期</label>
            <div class="segmented" id="trend-period"></div>
          </div>
        </div>
        <div class="chart-area" id="trend-chart"></div>
        <div class="table-wrap" id="trend-table"></div>
      </div>
    </section>

    <section class="view" id="view-returns">
      <div class="panel">
        <div class="panel-head">
          <h2>股票涨跌</h2>
          <p>按所选时间区间计算涨跌幅，并从高到低排序；股票使用全收益指数，商品使用可交易价格或商品指数。</p>
        </div>
        <div class="controls">
          <div class="control-wide">
            <label>计算区间</label>
            <div class="segmented" id="return-period"></div>
          </div>
          <div>
            <label for="return-start">起始日期</label>
            <input type="date" id="return-start">
          </div>
          <div>
            <label for="return-end">结束日期</label>
            <input type="date" id="return-end">
          </div>
        </div>
        <div class="chart-area" id="return-chart"></div>
        <div class="table-wrap" id="return-table"></div>
        <h3 style="margin-top:32px">中信风格 + 万得全A</h3>
        <div class="controls">
          <div class="control-wide">
            <label>计算区间</label>
            <div class="segmented" id="style-period"></div>
          </div>
          <div>
            <label for="style-start">起始日期</label>
            <input type="date" id="style-start">
          </div>
          <div>
            <label for="style-end">结束日期</label>
            <input type="date" id="style-end">
          </div>
        </div>
        <div class="chart-area" id="return-chart-style"></div>
        <div class="table-wrap" id="return-table-style"></div>
      </div>
    </section>

    <section class="view" id="view-rates">
      <div class="panel">
        <div class="panel-head">
          <h2>利率涨跌</h2>
          <p>按所选时间区间计算数据库中所有利率指标的 bp 变化，并从上行最多到下行最多排序。</p>
        </div>
        <div class="controls">
          <div class="control-wide">
            <label>计算区间</label>
            <div class="segmented" id="rate-period"></div>
          </div>
          <div>
            <label for="rate-start">起始日期</label>
            <input type="date" id="rate-start">
          </div>
          <div>
            <label for="rate-end">结束日期</label>
            <input type="date" id="rate-end">
          </div>
        </div>
        <div class="chart-area" id="rate-chart"></div>
        <div class="table-wrap" id="rate-table"></div>
      </div>
    </section>

    <section class="view" id="view-fx">
      <div class="panel">
        <div class="panel-head">
          <h2>汇率涨跌</h2>
          <p>按所选时间区间计算主要汇率和外汇指数的涨跌幅，并从高到低排序。</p>
        </div>
        <div class="controls">
          <div class="control-wide">
            <label>计算区间</label>
            <div class="segmented" id="fx-period"></div>
          </div>
          <div>
            <label for="fx-start">起始日期</label>
            <input type="date" id="fx-start">
          </div>
          <div>
            <label for="fx-end">结束日期</label>
            <input type="date" id="fx-end">
          </div>
        </div>
        <div class="chart-area" id="fx-chart"></div>
        <div class="table-wrap" id="fx-table"></div>
      </div>
    </section>

    <section class="view" id="view-spread">
      <div class="panel">
        <div class="panel-head">
          <h2>中美利差</h2>
          <p>展示 10Y / 2Y 美中利差，即美国收益率减中国收益率，并分别与 USDCNY 绘制在一张图上。</p>
        </div>
        <div class="controls">
          <div class="control-wide">
            <label>展示周期</label>
            <div class="segmented" id="spread-period"></div>
          </div>
        </div>
        <div class="chart-area" id="spread-chart-10y"></div>
        <div class="chart-area" id="spread-chart-2y"></div>
        <div class="table-wrap" id="spread-table"></div>
      </div>
    </section>

    <section class="view" id="view-valuation">
      <div class="panel">
        <div class="panel-head">
          <h2>估值看板</h2>
          <p>选择指数和估值指标，绘制尽可能长的历史序列，并叠加过去 N 年均值与 ±1 / ±2 倍标准差参考线。</p>
        </div>
        <div class="controls">
          <div>
            <label for="valuation-index">指数</label>
            <select id="valuation-index"></select>
          </div>
          <div>
            <label for="valuation-metric">估值指标</label>
            <select id="valuation-metric"></select>
          </div>
          <div class="control-wide">
            <label>N 年窗口</label>
            <div class="segmented" id="valuation-years"></div>
          </div>
          <div class="control-wide">
            <label>图表显示区间</label>
            <div class="segmented" id="valuation-display"></div>
          </div>
        </div>
        <div class="chart-area" id="valuation-chart"></div>
        <div class="table-wrap" id="valuation-table"></div>
      </div>
    </section>

    <section class="view" id="view-emotion">
      <div class="panel">
        <div class="panel-head">
          <h2>市场情绪</h2>
          <p>数据来源：华泰智研 MCP。情绪指数反映市场整体情绪状态（0-100），与对应市场基准指数对比展示。</p>
        </div>
        <div class="controls">
          <div class="control-wide">
            <label>展示周期</label>
            <div class="segmented" id="emotion-period"></div>
          </div>
        </div>
        <h3>A股 / 港股情绪指数</h3>
        <div class="chart-area" id="emotion-chart"></div>
        <div class="table-wrap" id="emotion-table"></div>
        <h3 style="margin-top:32px">A股情绪 vs 万得全A</h3>
        <div class="chart-area" id="emotion-chart-a"></div>
        <div class="table-wrap" id="emotion-table-a"></div>
        <h3 style="margin-top:32px">港股情绪 vs 恒生指数</h3>
        <div class="chart-area" id="emotion-chart-hk"></div>
        <div class="table-wrap" id="emotion-table-hk"></div>
      </div>
    </section>

    <section class="view" id="view-topics">
      <div class="panel">
        <div class="panel-head">
          <h2>专题图表</h2>
          <p>专题研究与定期更新的主题图表。当前专题：美元超级周期 — 将历史美元下行周期对齐到峰值月 (T+0)，归一化后叠加比较。</p>
        </div>
        <h3>Dollar Super Cycle Scenarios (DXY)</h3>
        <div class="topic-legend" id="topic-dxy-legend"></div>
        <div class="topic-chart-area" id="topic-chart-dxy"></div>
        <div class="table-wrap" id="topic-table-dxy"></div>
        <h3 style="margin-top:32px">Dollar Super Cycle Scenarios (Real Broad Dollar vs AE)</h3>
        <div class="topic-legend" id="topic-dae-legend"></div>
        <div class="topic-chart-area" id="topic-chart-dae"></div>
        <div class="table-wrap" id="topic-table-dae"></div>
      </div>
    </section>

    <section class="view" id="view-fixing">
      <div class="panel">
        <div class="panel-head">
          <h2>中间价</h2>
          <p>USDCNY 中间价 vs 即期汇率（±2% 波动区间）+ 即期汇率变动拆解（20 日滚动）。数据来源：CFETS。</p>
        </div>
        <div class="controls">
          <div class="control-wide">
            <label>展示周期</label>
            <div class="segmented" id="fixing-period"></div>
          </div>
        </div>
        <h3>中间价 vs 即期汇率</h3>
        <div class="chart-area" id="fixing-chart-1"></div>
        <div class="table-wrap" id="fixing-table-1"></div>
        <h3 style="margin-top:32px">即期汇率变动拆解（20 日滚动）</h3>
        <div class="chart-area" id="fixing-chart-2"></div>
        <div class="table-wrap" id="fixing-table-2"></div>
      </div>
    </section>

    <section class="view" id="view-cost">
      <div class="panel">
        <div class="panel-head">
          <h2>套保成本</h2>
          <p>CNY / CNH 远期套保成本期限结构与时序分析。套保成本为公式计算值，非外部 MCP 数据。</p>
        </div>

        <!-- Chart 1: Term Structure -->
        <h3>套保成本曲线</h3>
        <div class="controls" style="grid-template-columns: repeat(3, minmax(0, 1fr));">
          <div>
            <label>合约选择</label>
            <div class="segmented" id="cost-contract"></div>
          </div>
          <div>
            <label>取值方式（可多选）</label>
            <div class="mode-checks" id="cost-mode"></div>
          </div>
          <div id="cost-custom-wrap" style="display:none">
            <label>指定区间</label>
            <div style="display:flex; gap:8px; align-items:center">
              <input type="date" id="cost-range-start">
              <span style="color:var(--muted)">—</span>
              <input type="date" id="cost-range-end">
            </div>
          </div>
        </div>
        <div class="chart-area" id="cost-chart-1"></div>
        <div class="table-wrap" id="cost-table-1"></div>

        <!-- Chart 2: 3M Time Series -->
        <h3 style="margin-top:32px">3M 套保成本时序</h3>
        <div class="controls">
          <div class="control-wide">
            <label>展示周期</label>
            <div class="segmented" id="cost-3m-period"></div>
          </div>
          <div>
            <label>合约</label>
            <div class="mode-checks" id="cost-3m-contract"></div>
          </div>
          <div>
            <label>平滑</label>
            <div class="segmented" id="cost-3m-ma"></div>
          </div>
        </div>
        <div class="chart-area" id="cost-chart-2"></div>
        <div class="table-wrap" id="cost-table-2"></div>

        <!-- Chart 3: 1Y Hedge vs Spread -->
        <h3 style="margin-top:32px">1Y 套保成本 vs 中美利差</h3>
        <div class="controls">
          <div class="control-wide">
            <label>展示周期</label>
            <div class="segmented" id="cost-1y-period"></div>
          </div>
          <div>
            <label>合约</label>
            <div class="mode-checks" id="cost-1y-contract"></div>
          </div>
          <div>
            <label>平滑</label>
            <div class="segmented" id="cost-1y-ma"></div>
          </div>
        </div>
        <div class="chart-area" id="cost-chart-3"></div>
        <div class="table-wrap" id="cost-table-3"></div>
      </div>
    </section>

    <section class="panel">
      <div class="panel-head">
        <h2>数据口径</h2>
      </div>
      <div class="sources" id="sources"></div>
    </section>
  </main>

  <script id="dashboard-data" type="application/json">__DATA__</script>
  <script>
const DATA = JSON.parse(document.getElementById("dashboard-data").textContent);
const OBS = DATA.observations;
const META = DATA.meta;

const COLORS = {
  blue: "#5B7F8A",
  red: "#C44B3B",
  green: "#6B8B7A",
  gold: "#B8974A",
  purple: "#8B6B7E",
  muted: "#8B8581",
  grid: "#EAE4DA",
  axis: "#B8AFA4",
  A股: "#5B7F8A",
  A股风格: "#7DA0AA",
  港股: "#6B8B7A",
  美股: "#C44B3B",
  商品: "#B8974A",
  兑人民币: "#5B7F8A",
  美元交叉: "#6B8B7A",
  指数: "#B8974A"
};

const PERIODS = [
  ["15Y", "过去 15Y"], ["10Y", "过去 10Y"], ["5Y", "过去 5Y"], ["3Y", "过去 3Y"], ["1Y", "过去 1Y"],
  ["YTD", "YTD"], ["60D", "过去 60 日"], ["20D", "过去 20 日"], ["MARTIN", "Since 1989"]
];
const SPREAD_PERIODS = [["15Y", "过去 15Y"], ["10Y", "过去 10Y"], ["5Y", "5Y"], ["3Y", "3Y"], ["1Y", "1Y"], ["CNY2015", "Since 20150811"], ["CNY2005", "Since 20050721"]];
const SPREAD_SERIES = ["trend:中-10债", "trend:美-10债", "trend:中-2债", "trend:美-2债", "trend:USDCNY"];
const EMOTION_SERIES = ["htsc:A股情绪指数", "htsc:港股情绪指数"];

const RETURN_PERIODS = [["15Y", "过去 15Y"], ["10Y", "过去 10Y"], ["5Y", "过去 5Y"], ["3Y", "过去 3Y"], ["1Y", "过去一年"], ["YTD", "YTD"], ["QTD", "QTD"], ["CUSTOM", "自定义"]];
const RATE_FX_PERIODS = [["10Y", "过去 10Y"], ["5Y", "过去 5Y"], ["3Y", "过去 3Y"], ["1Y", "过去一年"], ["YTD", "YTD"], ["QTD", "QTD"], ["CUSTOM", "自定义"]];
const EMOTION_PERIODS = [["FULL", "全历史"], ["10Y", "过去 10Y"], ["5Y", "过去 5Y"], ["3Y", "过去 3Y"], ["1Y", "过去 1Y"], ["YTD", "YTD"]];
const VALUATION_WINDOWS = [["3", "3年"], ["5", "5年"], ["10", "10年"], ["15", "15年"], ["FULL", "全样本"]];
const VALUATION_DISPLAY_PERIODS = [["FULL", "全历史"], ["20Y", "过去 20Y"], ["15Y", "15年"], ["10Y", "10年"], ["5Y", "5年"], ["3Y", "3年"]];

function toDate(s) { return new Date(s + "T00:00:00"); }
function toISO(d) {
  const z = new Date(d.getTime() - d.getTimezoneOffset() * 60000);
  return z.toISOString().slice(0, 10);
}
function addYears(d, n) {
  const x = new Date(d);
  x.setFullYear(x.getFullYear() + n);
  return x;
}
function addDays(d, n) {
  const x = new Date(d);
  x.setDate(x.getDate() + n);
  return x;
}
function quarterStart(d) {
  const month = d.getMonth();
  const q = Math.floor(month / 3) * 3;
  return new Date(d.getFullYear(), q, 1);
}
function latestDate(seriesIds) {
  let latest = null;
  for (const sid of seriesIds) {
    const pts = OBS[sid] || [];
    if (!pts.length) continue;
    const d = pts[pts.length - 1][0];
    if (!latest || d > latest) latest = d;
  }
  return latest;
}
function startForPeriod(period, endDate) {
  const end = toDate(endDate);
  const yearMatch = String(period).match(/^(\d+)Y$/);
  if (yearMatch) return toISO(addYears(end, -Number(yearMatch[1])));
  if (period === "5Y") return toISO(addYears(end, -5));
  if (period === "3Y") return toISO(addYears(end, -3));
  if (period === "1Y") return toISO(addYears(end, -1));
  if (period === "YTD") return `${end.getFullYear()}-01-01`;
  if (period === "QTD") return toISO(quarterStart(end));
  if (period === "120D") return toISO(addDays(end, -120));
  if (period === "60D") return toISO(addDays(end, -60));
  if (period === "20D") return toISO(addDays(end, -20));
  if (period === "MARTIN") return "1989-06-05";
  if (period === "CNY2015") return "2015-08-11";
  if (period === "CNY2005") return "2005-07-21";
  return "1900-01-01";
}
function pointsInRange(sid, start, end, opts = {}) {
  const { skipZero = true } = opts;
  let pts = (OBS[sid] || []).filter(p => p[0] >= start && p[0] <= end);
  if (skipZero) pts = skipZeros(pts);
  return pts;
}
function nearestOnOrBefore(sid, target) {
  const pts = OBS[sid] || [];
  let best = null;
  for (const p of pts) {
    if (p[1] === 0 || !Number.isFinite(p[1])) continue;
    if (p[0] <= target) best = p;
    else break;
  }
  return best;
}
function axisFromValues(values) {
  const clean = values.filter(v => Number.isFinite(v));
  if (!clean.length) return [0, 1];
  const min = Math.min(...clean);
  const max = Math.max(...clean);
  if (max === min) {
    const pad = Math.max(Math.abs(max) * 0.1, 1);
    return [min - pad, max + pad];
  }
  const span = (max - min) / 0.7;
  return [min - span * 0.15, min + span * 0.85];
}
function formatValue(v, unit) {
  if (!Number.isFinite(v)) return "n/a";
  if (unit === "percent_point" || unit === "percent") return `${v.toFixed(2)}%`;
  if (unit === "fx") return v.toFixed(4);
  if (unit === "multiple") return `${v.toFixed(2)}x`;
  if (Math.abs(v) >= 1000) return v.toLocaleString("en-US", { maximumFractionDigits: 2 });
  return v.toFixed(2);
}
function formatPct(v) {
  if (!Number.isFinite(v)) return "n/a";
  return `${v >= 0 ? "+" : ""}${(v * 100).toFixed(2)}%`;
}
function formatBp(v) {
  if (!Number.isFinite(v)) return "n/a";
  return `${v >= 0 ? "+" : ""}${v.toFixed(0)}bp`;
}
function pctChange(series, start, end) {
  const a = nearestOnOrBefore(series, start);
  const b = nearestOnOrBefore(series, end);
  if (!a || !b || !a[1]) return null;
  // 数据不足：首个可用数据点晚于请求起点 → 跳过
  if (a[0] > start) return null;
  return { startDate: a[0], endDate: b[0], start: a[1], end: b[1], value: b[1] / a[1] - 1 };
}
function pctChangeCAGR(series, start, end) {
  const r = pctChange(series, start, end);
  if (!r) return null;
  const years = (toDate(r.endDate) - toDate(r.startDate)) / 365.25 / 86400000;
  if (years <= 1.05) return r;  // ≤1 年不年化
  const cagr = Math.pow(1 + r.value, 1 / years) - 1;
  return { ...r, value: cagr, annualized: true };
}
function bpChange(series, start, end) {
  const a = nearestOnOrBefore(series, start);
  const b = nearestOnOrBefore(series, end);
  if (!a || !b) return null;
  return { startDate: a[0], endDate: b[0], start: a[1], end: b[1], value: (b[1] - a[1]) * 100 };
}
function scale(value, min, max, low, high) {
  if (max === min) return (low + high) / 2;
  return low + (value - min) / (max - min) * (high - low);
}
function dateTicks(minX, maxX, count = 8) {
  if (maxX <= minX) return [minX];
  const ticks = [];
  for (let i = 0; i < count; i++) {
    ticks.push(minX + (maxX - minX) * i / (count - 1));
  }
  return ticks;
}
function formatDateTick(ms, minX, maxX) {
  const iso = new Date(ms).toISOString().slice(0, 10);
  const days = (maxX - minX) / 86400000;
  if (days > 1100) return iso.slice(0, 7);
  return iso;
}
function skipZeros(points) {
  return points.filter(p => p[1] !== 0 && Number.isFinite(p[1]));
}
function fillZeroWithPrevious(points) {
  const filled = [];
  let previous = null;
  for (const p of points) {
    const d = p[0], v = p[1];
    if (v == null || !Number.isFinite(v)) continue;
    if (v === 0) {
      if (previous == null) continue;
      filled.push([d, previous]);
    } else {
      previous = v;
      filled.push([d, v]);
    }
  }
  return filled;
}
function renderSegmented(id, options, active, onClick) {
  const el = document.getElementById(id);
  el.innerHTML = "";
  for (const [value, label] of options) {
    const btn = document.createElement("button");
    btn.type = "button";
    btn.textContent = label;
    btn.dataset.value = value;
    btn.className = value === active ? "active" : "";
    btn.addEventListener("click", () => {
      [...el.querySelectorAll("button")].forEach(b => b.classList.remove("active"));
      btn.classList.add("active");
      onClick(value);
    });
    el.appendChild(btn);
  }
}
function renderSvgLine(containerId, cfg) {
  const width = 980, height = 420;
  const left = 64, right = cfg.rightAxis ? 64 : 22, top = 62, bottom = 58;
  const plotW = width - left - right, plotH = height - top - bottom;
  const allX = cfg.series.flatMap(s => s.points.map(p => toDate(p[0]).getTime()));
  if (!allX.length) {
    document.getElementById(containerId).innerHTML = '<div class="empty">暂无可用数据</div>';
    return;
  }
  const minX = Math.min(...allX), maxX = Math.max(...allX);
  const yLeft = cfg.leftAxis;
  const yRight = cfg.rightAxis || cfg.leftAxis;
  const grid = [];
  for (let i = 0; i <= 4; i++) {
    const y = top + i * plotH / 4;
    const valLeft = yLeft[1] - i * (yLeft[1] - yLeft[0]) / 4;
    grid.push(`<line x1="${left}" y1="${y.toFixed(1)}" x2="${width - right}" y2="${y.toFixed(1)}" class="grid-line"/>`);
    grid.push(`<text x="${left - 10}" y="${(y + 4).toFixed(1)}" text-anchor="end" class="axis-text">${cfg.leftFormat(valLeft)}</text>`);
    if (cfg.rightAxis) {
      const valRight = yRight[1] - i * (yRight[1] - yRight[0]) / 4;
      grid.push(`<text x="${width - right + 10}" y="${(y + 4).toFixed(1)}" class="axis-text">${cfg.rightFormat(valRight)}</text>`);
    }
  }
  for (const tick of dateTicks(minX, maxX, 8)) {
    const x = scale(tick, minX, maxX, left, left + plotW);
    grid.push(`<line x1="${x.toFixed(1)}" y1="${top}" x2="${x.toFixed(1)}" y2="${top + plotH}" class="grid-line"/>`);
    grid.push(`<text x="${x.toFixed(1)}" y="${height - 30}" text-anchor="middle" class="axis-text">${formatDateTick(tick, minX, maxX)}</text>`);
  }
  const paths = [];
  const legends = [];
  const markers = [];
  const hitTargets = [];
  cfg.series.forEach((s, idx) => {
    const axis = s.axis === "right" ? yRight : yLeft;
    const fmt = s.axis === "right" ? cfg.rightFormat : cfg.leftFormat;
    const d = s.points.map((p, i) => {
      const x = scale(toDate(p[0]).getTime(), minX, maxX, left, left + plotW);
      const y = scale(p[1], axis[0], axis[1], top + plotH, top);
      return `${i === 0 ? "M" : "L"}${x.toFixed(1)},${y.toFixed(1)}`;
    }).join(" ");
    const dash = s.dash ? ` stroke-dasharray="${s.dash}"` : "";
    paths.push(`<path d="${d}" fill="none" stroke="${s.color}" stroke-width="${s.width || 2.2}"${dash}/>` );
    if ((s.markers ?? true) && s.points.length <= 320) {
      for (const p of s.points) {
        const x = scale(toDate(p[0]).getTime(), minX, maxX, left, left + plotW);
        const y = scale(p[1], axis[0], axis[1], top + plotH, top);
        markers.push(`<circle cx="${x.toFixed(1)}" cy="${y.toFixed(1)}" r="1.7" fill="${s.color}" opacity="0.85"/>`);
      }
    }
    for (const p of s.points) {
      const x = scale(toDate(p[0]).getTime(), minX, maxX, left, left + plotW);
      const y = scale(p[1], axis[0], axis[1], top + plotH, top);
      const label = `${s.name}|${p[0]}|${fmt(p[1])}`;
      hitTargets.push(`<circle class="chart-hit" cx="${x.toFixed(1)}" cy="${y.toFixed(1)}" r="6" data-color="${s.color}" data-label="${label}"><title>${label}</title></circle>`);
    }
    const lx = left + (idx % 4) * 210;
    const ly = 24 + Math.floor(idx / 4) * 20;
    legends.push(`<line x1="${lx}" y1="${ly - 4}" x2="${lx + 18}" y2="${ly - 4}" stroke="${s.color}" stroke-width="2.8"${dash}/><text x="${lx + 24}" y="${ly}" class="legend-text">${s.name}</text>`);
  });
  const html = `<svg class="chart" viewBox="0 0 ${width} ${height}" role="img" aria-label="${cfg.title}">
    ${grid.join("")}
    <line x1="${left}" y1="${top + plotH}" x2="${width - right}" y2="${top + plotH}" class="axis-line"/>
    <line x1="${left}" y1="${top}" x2="${left}" y2="${top + plotH}" class="axis-line"/>
    ${cfg.rightAxis ? `<line x1="${width - right}" y1="${top}" x2="${width - right}" y2="${top + plotH}" class="axis-line"/>` : ""}
    ${paths.join("")}
    ${markers.join("")}
    ${hitTargets.join("")}
    <g class="chart-tooltip" style="display:none">
      <rect class="tooltip-box" x="0" y="0" width="178" height="54" rx="6"/>
      <text class="tooltip-title" x="10" y="20"></text>
      <text class="tooltip-text" x="10" y="40"></text>
    </g>
    ${legends.join("")}
  </svg>`;
  const container = document.getElementById(containerId);
  container.innerHTML = html;
  const svg = container.querySelector("svg");
  const tooltip = svg.querySelector(".chart-tooltip");
  const tooltipTitle = tooltip.querySelector(".tooltip-title");
  const tooltipText = tooltip.querySelector(".tooltip-text");
  const tooltipRect = tooltip.querySelector(".tooltip-box");
  const tipW = 178, tipH = 54;
  function showPointTip(target) {
    const parts = target.dataset.label.split("|");
    const x = Number(target.getAttribute("cx"));
    const y = Number(target.getAttribute("cy"));
    const tx = Math.min(Math.max(x + 12, left + 4), left + plotW - tipW - 4);
    const ty = Math.min(Math.max(y - tipH - 12, top + 4), top + plotH - tipH - 4);
    tooltip.setAttribute("transform", `translate(${tx.toFixed(1)} ${ty.toFixed(1)})`);
    tooltipRect.setAttribute("stroke", target.dataset.color);
    tooltipTitle.textContent = parts[0];
    tooltipText.textContent = `${parts[1]}  ${parts[2]}`;
    tooltip.style.display = "";
  }
  svg.querySelectorAll(".chart-hit").forEach(hit => {
    hit.addEventListener("mouseenter", event => showPointTip(event.currentTarget));
    hit.addEventListener("mousemove", event => showPointTip(event.currentTarget));
    hit.addEventListener("click", event => showPointTip(event.currentTarget));
  });
  svg.addEventListener("mouseleave", () => { tooltip.style.display = "none"; });
}
function tableHtml(headers, rows) {
  return `<table><thead><tr>${headers.map(h => `<th>${h}</th>`).join("")}</tr></thead><tbody>${rows.map(r => `<tr>${r.map(c => `<td>${c}</td>`).join("")}</tr>`).join("")}</tbody></table>`;
}

let trendPeriod = "1Y";
let returnPeriod = "1Y";
let stylePeriod = "1Y";
let ratePeriod = "1Y";
let fxPeriod = "1Y";
let spreadPeriod = "3Y";
let valuationWindow = "10";
let valuationDisplayPeriod = "FULL";
let emotionPeriod = "1Y";

function initTrend() {
  const one = document.getElementById("trend-one");
  const two = document.getElementById("trend-two");
  one.innerHTML = "";
  two.innerHTML = `<option value="">不选择第二个指标</option>`;
  DATA.trend_options.forEach(opt => {
    const label = `${opt.display_name}`;
    one.add(new Option(label, opt.series_id));
    two.add(new Option(label, opt.series_id));
  });
  one.value = "trend:美-10债";
  two.value = "trend:USDCNY";
  one.addEventListener("change", renderTrend);
  two.addEventListener("change", renderTrend);
  renderSegmented("trend-period", PERIODS, trendPeriod, v => { trendPeriod = v; renderTrend(); });
  renderTrend();
}
function renderTrend() {
  const s1 = document.getElementById("trend-one").value;
  const s2 = document.getElementById("trend-two").value;
  const dual = s2 && s2 !== s1;
  const end = latestDate([s1, s2].filter(Boolean));
  const start = startForPeriod(trendPeriod, end);
  // 单指标：跳过零值；双指标：先不过滤零值，再前向填补保证两条序列时间轴对齐
  const raw1 = pointsInRange(s1, start, end, { skipZero: !dual });
  const raw2 = dual ? pointsInRange(s2, start, end, { skipZero: false }) : [];
  const p1 = dual ? fillZeroWithPrevious(raw1) : raw1;
  const p2 = dual ? fillZeroWithPrevious(raw2) : raw2;
  const unit1 = META[s1].unit;
  const axis1 = axisFromValues(p1.map(p => p[1]));
  const axis2 = p2.length ? axisFromValues(p2.map(p => p[1])) : null;
  renderSvgLine("trend-chart", {
    title: "走势看板",
    leftAxis: axis1,
    rightAxis: axis2 || axis1,
    leftFormat: v => formatValue(v, unit1),
    rightFormat: v => formatValue(v, p2.length ? META[s2].unit : unit1),
    series: [
      { name: META[s1].display_name, points: p1, color: COLORS.blue, axis: "left" },
      ...(p2.length ? [{ name: META[s2].display_name, points: p2, color: COLORS.red, axis: "right" }] : [])
    ]
  });
  const rows = [s1, ...(p2.length ? [s2] : [])].map(sid => {
    const ptsForRow = dual
      ? fillZeroWithPrevious(pointsInRange(sid, start, end, { skipZero: false }))
      : pointsInRange(sid, start, end);
    const first = ptsForRow[0], last = ptsForRow[ptsForRow.length - 1];
    let chg = "n/a";
    if (first && last) {
      chg = META[sid].unit === "percent_point"
        ? formatBp((last[1] - first[1]) * 100)
        : (first[1] ? formatPct(last[1] / first[1] - 1) : "n/a");
    }
    return [META[sid].display_name, first?.[0] || "n/a", last?.[0] || "n/a", ptsForRow.length, formatValue(last?.[1], META[sid].unit), chg];
  });
  document.getElementById("trend-table").innerHTML = tableHtml(["指标", "起点", "终点", "样本点数", "最新值", "区间涨跌"], rows);
}

function initReturns() {
  const end = latestDate(DATA.return_options.map(o => o.series_id));
  document.getElementById("return-end").value = end;
  document.getElementById("return-start").value = startForPeriod(returnPeriod, end);
  document.getElementById("style-end").value = end;
  document.getElementById("style-start").value = startForPeriod(stylePeriod, end);
  document.getElementById("return-start").addEventListener("change", () => { returnPeriod = "CUSTOM"; activateSegment("return-period", "CUSTOM"); renderReturns(); });
  document.getElementById("return-end").addEventListener("change", () => { returnPeriod = "CUSTOM"; activateSegment("return-period", "CUSTOM"); renderReturns(); });
  document.getElementById("style-start").addEventListener("change", () => { stylePeriod = "CUSTOM"; activateSegment("style-period", "CUSTOM"); renderReturns(); });
  document.getElementById("style-end").addEventListener("change", () => { stylePeriod = "CUSTOM"; activateSegment("style-period", "CUSTOM"); renderReturns(); });
  renderSegmented("return-period", RETURN_PERIODS, returnPeriod, v => {
    returnPeriod = v;
    if (!stylePeriod || stylePeriod === returnPeriod) { stylePeriod = v; }
    if (v !== "CUSTOM") {
      const e = latestDate(DATA.return_options.map(o => o.series_id));
      document.getElementById("return-end").value = e;
      document.getElementById("return-start").value = startForPeriod(v, e);
    }
    renderReturns();
  });
  renderSegmented("style-period", RETURN_PERIODS, stylePeriod, v => {
    stylePeriod = v;
    if (v !== "CUSTOM") {
      const e = latestDate(DATA.return_options.map(o => o.series_id));
      document.getElementById("style-end").value = e;
      document.getElementById("style-start").value = startForPeriod(v, e);
    }
    renderReturns();
  });
  renderReturns();
}
function activateSegment(id, value) {
  [...document.getElementById(id).querySelectorAll("button")].forEach(b => b.classList.toggle("active", b.dataset.value === value));
}
function renderBarChart(containerId, values, cfg) {
  const maxAbs = Math.max(...values.map(v => Math.abs(v.value)), cfg.minAbs || 0.01);
  const width = 980, barH = 26, gap = 10, left = cfg.left || 185, right = 90, top = 48, bottom = 34;
  const height = top + bottom + values.length * (barH + gap);
  const zeroX = left + (width - left - right) / 2;
  const scaleX = (width - left - right) / 2 / maxAbs;
  const legendItems = cfg.legend || [];
  const legend = legendItems.map((item, idx) => {
    const x = left + idx * 120;
    return `<rect x="${x}" y="14" width="12" height="12" fill="${item.color}"/><text x="${x + 18}" y="25" class="legend-text">${item.label}</text>`;
  }).join("");
  const ticks = [-maxAbs, -maxAbs / 2, 0, maxAbs / 2, maxAbs];
  const axisTicks = ticks.map(t => {
    const x = zeroX + t * scaleX;
    return `<line x1="${x}" y1="${top - 8}" x2="${x}" y2="${height - bottom + 8}" class="grid-line"/>
      <text x="${x}" y="${height - 12}" text-anchor="middle" class="axis-text">${cfg.format(t)}</text>`;
  }).join("");
  const bars = values.map((v, i) => {
    const y = top + i * (barH + gap);
    const w = Math.max(Math.abs(v.value) * scaleX, 2.5);
    const x = v.value >= 0 ? zeroX : zeroX - w;
    const outsideX = v.value >= 0 ? x + w + 8 : x - 8;
    const outsideAnchor = v.value >= 0 ? "start" : "end";
    const nearLeftAxis = outsideX < left + 8;
    const nearRightEdge = outsideX > width - right - 8;
    const labelX = nearLeftAxis ? x + w + 8 : nearRightEdge ? x - 8 : outsideX;
    const anchor = nearLeftAxis ? "start" : nearRightEdge ? "end" : outsideAnchor;
    const fill = nearLeftAxis || nearRightEdge ? "" : "";
    return `<text x="${left - 18}" y="${y + 18}" text-anchor="end" class="axis-text">${v.display_name}</text>
      <rect x="${x}" y="${y}" width="${Math.max(w, 1)}" height="${barH}" fill="${v.color}"/>
      <text x="${labelX}" y="${y + 18}" text-anchor="${anchor}" class="legend-text"${fill}>${cfg.format(v.value)}</text>`;
  }).join("");
  document.getElementById(containerId).style.minHeight = "";
  document.getElementById(containerId).innerHTML = `<svg class="chart" viewBox="0 0 ${width} ${height}" style="height:${height}px;min-width:${width}px" role="img" aria-label="${cfg.title}">
    ${legend}
    ${axisTicks}
    <line x1="${zeroX}" y1="${top - 8}" x2="${zeroX}" y2="${height - bottom + 8}" class="axis-line"/>
    ${bars}
  </svg>`;
}
function renderReturns() {
  // Chart 1: 股票与商品
  const mainStart = document.getElementById("return-start").value;
  const mainEnd = document.getElementById("return-end").value;
  const isAnnualMain = returnPeriod.match(/^\d+Y$/) && Number(returnPeriod.match(/^(\d+)Y$/)[1]) > 1;
  const mainRegions = ["A股", "港股", "美股", "商品"];
  const mainValues = DATA.return_options
    .filter(opt => mainRegions.includes(opt.region))
    .map(opt => {
      const result = isAnnualMain ? pctChangeCAGR(opt.series_id, mainStart, mainEnd) : pctChange(opt.series_id, mainStart, mainEnd);
      return result ? { ...opt, ...result, color: COLORS[opt.region] } : null;
    }).filter(Boolean).sort((a, b) => b.value - a.value);
  const mainLabel = isAnnualMain ? "复合年化收益率" : "涨跌幅";
  renderBarChart("return-chart", mainValues, {
    title: `股票与商品${mainLabel}`,
    format: formatPct,
    legend: [
      { label: "A股", color: COLORS.A股 },
      { label: "港股", color: COLORS.港股 },
      { label: "美股", color: COLORS.美股 },
      { label: "商品", color: COLORS.商品 }
    ]
  });
  const mainRows = mainValues.map(v => [v.display_name, v.region, v.startDate, v.endDate, formatValue(v.start, META[v.series_id].unit), formatValue(v.end, META[v.series_id].unit), formatPct(v.value)]);
  document.getElementById("return-table").innerHTML = tableHtml(["标的", "类别", "起始读数日", "结束读数日", "起始值", "结束值", mainLabel], mainRows);

  // Chart 2: 中信风格 + 万得全A (独立区间)
  const styleStart = document.getElementById("style-start").value;
  const styleEnd = document.getElementById("style-end").value;
  const isAnnualStyle = stylePeriod.match(/^\d+Y$/) && Number(stylePeriod.match(/^(\d+)Y$/)[1]) > 1;
  const styleValues = DATA.return_options
    .filter(opt => opt.region === "A股风格")
    .map(opt => {
      const result = isAnnualStyle ? pctChangeCAGR(opt.series_id, styleStart, styleEnd) : pctChange(opt.series_id, styleStart, styleEnd);
      return result ? { ...opt, ...result, color: COLORS[opt.region] } : null;
    }).filter(Boolean).sort((a, b) => b.value - a.value);
  const styleLabel = isAnnualStyle ? "复合年化收益率" : "涨跌幅";
  renderBarChart("return-chart-style", styleValues, {
    title: `中信风格与万得全A${styleLabel}`,
    format: formatPct,
    legend: [
      { label: "中信风格/万得全A", color: COLORS.A股风格 }
    ]
  });
  const styleRows = styleValues.map(v => [v.display_name, "A股风格", v.startDate, v.endDate, formatValue(v.start, META[v.series_id].unit), formatValue(v.end, META[v.series_id].unit), formatPct(v.value)]);
  document.getElementById("return-table-style").innerHTML = tableHtml(["标的", "类别", "起始读数日", "结束读数日", "起始值", "结束值", styleLabel], styleRows);
}

function initRates() {
  const end = latestDate(DATA.rate_options.map(o => o.series_id));
  document.getElementById("rate-end").value = end;
  document.getElementById("rate-start").value = startForPeriod(ratePeriod, end);
  document.getElementById("rate-start").addEventListener("change", () => { ratePeriod = "CUSTOM"; activateSegment("rate-period", "CUSTOM"); renderRates(); });
  document.getElementById("rate-end").addEventListener("change", () => { ratePeriod = "CUSTOM"; activateSegment("rate-period", "CUSTOM"); renderRates(); });
  renderSegmented("rate-period", RATE_FX_PERIODS, ratePeriod, v => {
    ratePeriod = v;
    if (v !== "CUSTOM") {
      const e = latestDate(DATA.rate_options.map(o => o.series_id));
      document.getElementById("rate-end").value = e;
      document.getElementById("rate-start").value = startForPeriod(v, e);
    }
    renderRates();
  });
  renderRates();
}
function renderRates() {
  const start = document.getElementById("rate-start").value;
  const end = document.getElementById("rate-end").value;
  const isLong = ratePeriod.match(/^\d+Y$/) && Number(ratePeriod.match(/^(\d+)Y$/)[1]) > 1;
  const values = DATA.rate_options.map(opt => {
    const result = bpChange(opt.series_id, start, end);
    // 数据不足跳过
    if (result && result.startDate > start) return null;
    return result ? { ...opt, ...result, color: result.value >= 0 ? COLORS.red : COLORS.green } : null;
  }).filter(Boolean).sort((a, b) => b.value - a.value);
  const label = isLong ? "累计bp变化" : "bp变化";
  renderBarChart("rate-chart", values, {
    title: "利率涨跌柱状图",
    format: formatBp,
    minAbs: 10,
    legend: [
      { label: "利率上行", color: COLORS.red },
      { label: "利率下行", color: COLORS.green }
    ]
  });
  const rows = values.map(v => [v.display_name, v.startDate, v.endDate, formatValue(v.start, META[v.series_id].unit), formatValue(v.end, META[v.series_id].unit), formatBp(v.value)]);
  document.getElementById("rate-table").innerHTML = tableHtml(["利率", "起始读数日", "结束读数日", "起始值", "结束值", label], rows);
}

function initFx() {
  const end = latestDate(DATA.fx_options.map(o => o.series_id));
  document.getElementById("fx-end").value = end;
  document.getElementById("fx-start").value = startForPeriod(fxPeriod, end);
  document.getElementById("fx-start").addEventListener("change", () => { fxPeriod = "CUSTOM"; activateSegment("fx-period", "CUSTOM"); renderFx(); });
  document.getElementById("fx-end").addEventListener("change", () => { fxPeriod = "CUSTOM"; activateSegment("fx-period", "CUSTOM"); renderFx(); });
  renderSegmented("fx-period", RATE_FX_PERIODS, fxPeriod, v => {
    fxPeriod = v;
    if (v !== "CUSTOM") {
      const e = latestDate(DATA.fx_options.map(o => o.series_id));
      document.getElementById("fx-end").value = e;
      document.getElementById("fx-start").value = startForPeriod(v, e);
    }
    renderFx();
  });
  renderFx();
}
function renderFx() {
  const start = document.getElementById("fx-start").value;
  const end = document.getElementById("fx-end").value;
  const isLong = fxPeriod.match(/^\d+Y$/) && Number(fxPeriod.match(/^(\d+)Y$/)[1]) > 1;
  const values = DATA.fx_options.map(opt => {
    const result = pctChange(opt.series_id, start, end);
    if (result && result.startDate > start) return null;
    return result ? { ...opt, ...result, color: COLORS[opt.group] } : null;
  }).filter(Boolean).sort((a, b) => b.value - a.value);
  const label = isLong ? "累计涨跌幅" : "涨跌幅";
  renderBarChart("fx-chart", values, {
    title: "汇率涨跌柱状图",
    format: formatPct,
    legend: [
      { label: "兑人民币", color: COLORS.兑人民币 },
      { label: "美元交叉", color: COLORS.美元交叉 },
      { label: "指数", color: COLORS.指数 }
    ]
  });
  const rows = values.map(v => [v.display_name, v.group, v.startDate, v.endDate, formatValue(v.start, META[v.series_id].unit), formatValue(v.end, META[v.series_id].unit), formatPct(v.value)]);
  document.getElementById("fx-table").innerHTML = tableHtml(["标的", "类别", "起始读数日", "结束读数日", "起始值", "结束值", label], rows);
}

function spreadPoints(cnSid, usSid, start, end) {
  // 中美利差始终为双指标场景：先对原始序列前向填补，再计算利差
  const cnRaw = pointsInRange(cnSid, start, end, { skipZero: false });
  const usRaw = pointsInRange(usSid, start, end, { skipZero: false });
  const cn = fillZeroWithPrevious(cnRaw);
  const usMap = new Map(fillZeroWithPrevious(usRaw).map(p => [p[0], p[1]]));
  return cn.filter(p => usMap.has(p[0])).map(p => [p[0], (usMap.get(p[0]) - p[1]) * 100]);
}
function initSpread() {
  renderSegmented("spread-period", SPREAD_PERIODS, spreadPeriod, v => { spreadPeriod = v; renderSpread(); });
  renderSpread();
}
function renderSpread() {
  const end = latestDate(SPREAD_SERIES);
  const start = startForPeriod(spreadPeriod, end);
  // 双指标场景：USDCNY 也需要前向填补
  const fxRaw = pointsInRange("trend:USDCNY", start, end, { skipZero: false });
  const fx = fillZeroWithPrevious(fxRaw);
  const p10 = spreadPoints("trend:中-10债", "trend:美-10债", start, end);
  const p2 = spreadPoints("trend:中-2债", "trend:美-2债", start, end);
  renderSvgLine("spread-chart-10y", {
    title: "美中10Y利差与USDCNY",
    leftAxis: axisFromValues(p10.map(p => p[1])),
    rightAxis: axisFromValues(fx.map(p => p[1])),
    leftFormat: v => `${v.toFixed(0)}bp`,
    rightFormat: v => v.toFixed(4),
    series: [
      { name: "美中10Y利差", points: p10, color: COLORS.blue, axis: "left" },
      { name: "USDCNY", points: fx, color: COLORS.red, axis: "right" }
    ]
  });
  renderSvgLine("spread-chart-2y", {
    title: "美中2Y利差与USDCNY",
    leftAxis: axisFromValues(p2.map(p => p[1])),
    rightAxis: axisFromValues(fx.map(p => p[1])),
    leftFormat: v => `${v.toFixed(0)}bp`,
    rightFormat: v => v.toFixed(4),
    series: [
      { name: "美中2Y利差", points: p2, color: COLORS.green, axis: "left" },
      { name: "USDCNY", points: fx, color: COLORS.red, axis: "right" }
    ]
  });
  const rows = [
    ["美中10Y利差", p10[0]?.[0] || "n/a", p10[p10.length - 1]?.[0] || "n/a", p10.length, p10.length ? formatBp(p10[p10.length - 1][1]) : "n/a"],
    ["美中2Y利差", p2[0]?.[0] || "n/a", p2[p2.length - 1]?.[0] || "n/a", p2.length, p2.length ? formatBp(p2[p2.length - 1][1]) : "n/a"],
    ["USDCNY", fx[0]?.[0] || "n/a", fx[fx.length - 1]?.[0] || "n/a", fx.length, fx.length ? formatValue(fx[fx.length - 1][1], "fx") : "n/a"]
  ];
  document.getElementById("spread-table").innerHTML = tableHtml(["序列", "起点", "终点", "样本点数", "最新值"], rows);
}

function initValuation() {
  const idx = document.getElementById("valuation-index");
  const metric = document.getElementById("valuation-metric");
  DATA.valuation_options.forEach(opt => idx.add(new Option(opt.display_name, opt.display_name)));
  DATA.valuation_metrics.forEach(m => metric.add(new Option(m, m)));
  idx.value = "300收益";
  metric.value = "PE TTM";
  idx.addEventListener("change", renderValuation);
  metric.addEventListener("change", renderValuation);
  renderSegmented("valuation-years", VALUATION_WINDOWS, valuationWindow, v => { valuationWindow = v; renderValuation(); });
  renderSegmented("valuation-display", VALUATION_DISPLAY_PERIODS, valuationDisplayPeriod, v => { valuationDisplayPeriod = v; renderValuation(); });
  renderValuation();
}
function rollingReferenceSeries(points, window) {
  const means = [], up1 = [], down1 = [], up2 = [], down2 = [];
  for (let i = 0; i < points.length; i++) {
    const d = points[i][0];
    const start = window === "FULL" ? points[0][0] : startForPeriod(`${window}Y`, d);
    const vals = points.slice(0, i).filter(p => p[0] >= start && p[0] < d).map(p => p[1]);
    if (vals.length < 3) continue;
    const mean = vals.reduce((a, b) => a + b, 0) / vals.length;
    const variance = vals.reduce((a, b) => a + Math.pow(b - mean, 2), 0) / Math.max(vals.length - 1, 1);
    const sd = Math.sqrt(variance);
    means.push([d, mean]);
    up1.push([d, mean + sd]);
    down1.push([d, mean - sd]);
    up2.push([d, mean + 2 * sd]);
    down2.push([d, mean - 2 * sd]);
  }
  return { means, up1, down1, up2, down2 };
}
function latestWindowStats(points, years) {
  const latest = points[points.length - 1];
  const start = startForPeriod(`${years}Y`, latest[0]);
  const vals = points.filter(p => p[0] >= start && p[0] < latest[0]).map(p => p[1]);
  if (vals.length < 3) return null;
  const mean = vals.reduce((a, b) => a + b, 0) / vals.length;
  const variance = vals.reduce((a, b) => a + Math.pow(b - mean, 2), 0) / Math.max(vals.length - 1, 1);
  const sd = Math.sqrt(variance);
  return { mean, sd, z: sd === 0 ? null : (latest[1] - mean) / sd, count: vals.length };
}
function formatZ(z) {
  if (!Number.isFinite(z)) return "n/a";
  if (z > 0) return `高于均值 ${z.toFixed(2)}σ`;
  if (z < 0) return `低于均值 ${Math.abs(z).toFixed(2)}σ`;
  return "等于均值";
}
function renderValuation() {
  const name = document.getElementById("valuation-index").value;
  const metric = document.getElementById("valuation-metric").value;
  const item = DATA.valuation_options.find(x => x.display_name === name);
  const sid = item.metrics[metric];
  // 估值看板为单指标视图：跳过零值，视为当日无数据
  const rawPts = OBS[sid] || [];
  const pts = skipZeros(rawPts);
  if (!pts.length) return;
  const refs = rollingReferenceSeries(pts, valuationWindow);
  const end = pts[pts.length - 1][0];
  const displayStart = valuationDisplayPeriod === "FULL" ? pts[0][0] : startForPeriod(valuationDisplayPeriod, end);
  const displayPts = pts.filter(p => p[0] >= displayStart && p[0] <= end);
  const displayRefs = {
    means: refs.means.filter(p => p[0] >= displayStart && p[0] <= end),
    up1: refs.up1.filter(p => p[0] >= displayStart && p[0] <= end),
    down1: refs.down1.filter(p => p[0] >= displayStart && p[0] <= end),
    up2: refs.up2.filter(p => p[0] >= displayStart && p[0] <= end),
    down2: refs.down2.filter(p => p[0] >= displayStart && p[0] <= end)
  };
  const allValues = displayPts.map(p => p[1])
    .concat(displayRefs.means.map(p => p[1]), displayRefs.up1.map(p => p[1]), displayRefs.down1.map(p => p[1]), displayRefs.up2.map(p => p[1]), displayRefs.down2.map(p => p[1]));
  const axis = axisFromValues(allValues);
  const windowLabel = valuationWindow === "FULL" ? "全样本" : `${valuationWindow}年`;
  const displayLabel = valuationDisplayPeriod === "FULL" ? "全历史" : VALUATION_DISPLAY_PERIODS.find(x => x[0] === valuationDisplayPeriod)[1];
  renderSvgLine("valuation-chart", {
    title: `${name} ${metric}`,
    leftAxis: axis,
    leftFormat: v => formatValue(v, META[sid].unit),
    series: [
      { name: `${name} ${metric}`, points: displayPts, color: "#C47B5A", axis: "left" },
      { name: `${windowLabel}均值`, points: displayRefs.means, color: "#2C2416", width: 2.3, markers: false },
      { name: "+1σ", points: displayRefs.up1, color: "#8B8581", dash: "7 5", markers: false },
      { name: "-1σ", points: displayRefs.down1, color: "#8B8581", dash: "7 5", markers: false },
      { name: "+2σ", points: displayRefs.up2, color: "#A0988E", dash: "3 7", markers: false },
      { name: "-2σ", points: displayRefs.down2, color: "#A0988E", dash: "3 7", markers: false }
    ]
  });
  const latest = pts[pts.length - 1];
  const selectedStats = valuationWindow === "FULL" ? latestWindowStats(pts, 1000) : latestWindowStats(pts, Number(valuationWindow));
  const zRows = [5, 10, 15].map(years => {
    const stats = latestWindowStats(pts, years);
    return [
      `相对过去${years}年`,
      stats ? formatValue(stats.mean, META[sid].unit) : "n/a",
      stats ? formatValue(stats.sd, META[sid].unit) : "n/a",
      stats ? formatZ(stats.z) : "n/a"
    ];
  });
  const rows = [
    ["最新日期", latest[0], "", ""],
    ["最新估值", formatValue(latest[1], META[sid].unit), "", ""],
    [`当前图中参考窗口`, windowLabel, "", ""],
    [`当前图表显示区间`, displayLabel, displayPts[0]?.[0] || "n/a", displayPts.length],
    [`${windowLabel}最新均值`, selectedStats ? formatValue(selectedStats.mean, META[sid].unit) : "n/a", selectedStats ? formatValue(selectedStats.sd, META[sid].unit) : "n/a", selectedStats ? formatZ(selectedStats.z) : "n/a"],
    ...zRows
  ];
  document.getElementById("valuation-table").innerHTML = tableHtml(["项目", "均值/数值", "标准差", "当前位置"], rows);
}

function initEmotion() {
  renderSegmented("emotion-period", EMOTION_PERIODS, emotionPeriod, v => { emotionPeriod = v; renderEmotion(); });
  renderEmotion();
}
function renderEmotion() {
  const end = latestDate(EMOTION_SERIES);
  const start = emotionPeriod === "FULL" ? "1900-01-01" : startForPeriod(emotionPeriod, end);

  const aEmo = pointsInRange("htsc:A股情绪指数", start, end, { skipZero: true });
  const hkEmo = pointsInRange("htsc:港股情绪指数", start, end, { skipZero: true });

  // Chart 1: A股 + 港股情绪指数 (combined, single left axis)
  const allVals = [...aEmo.map(p => p[1]), ...hkEmo.map(p => p[1])];
  renderSvgLine("emotion-chart", {
    title: "市场情绪指数",
    leftAxis: axisFromValues(allVals),
    leftFormat: v => v.toFixed(1),
    series: [
      { name: "A股情绪", points: aEmo, color: COLORS.blue, axis: "left" },
      { name: "港股情绪", points: hkEmo, color: "#C47B5A", axis: "left" }
    ]
  });
  const rows = [
    ["A股情绪指数", aEmo[0]?.[0] || "n/a", aEmo[aEmo.length - 1]?.[0] || "n/a", aEmo.length,
     aEmo.length ? formatValue(aEmo[aEmo.length - 1][1], "percent") : "n/a"],
    ["港股情绪指数", hkEmo[0]?.[0] || "n/a", hkEmo[hkEmo.length - 1]?.[0] || "n/a", hkEmo.length,
     hkEmo.length ? formatValue(hkEmo[hkEmo.length - 1][1], "percent") : "n/a"]
  ];
  document.getElementById("emotion-table").innerHTML = tableHtml(["指标", "起点", "终点", "样本点数", "最新值"], rows);

  // Chart 2: A股情绪 vs 万得全A (dual axes)
  const wzqa = pointsInRange("trend:万得全A", start, end, { skipZero: true });
  if (aEmo.length || wzqa.length) {
    renderSvgLine("emotion-chart-a", {
      title: "A股情绪与万得全A",
      leftAxis: axisFromValues(aEmo.map(p => p[1])),
      rightAxis: axisFromValues(wzqa.map(p => p[1])),
      leftFormat: v => v.toFixed(1),
      rightFormat: v => formatValue(v, "index"),
      series: [
        { name: "A股情绪", points: aEmo, color: COLORS.blue, axis: "left" },
        { name: "万得全A", points: wzqa, color: "#C47B5A", axis: "right" }
      ]
    });
    const rowsA = [
      ["A股情绪指数", aEmo[0]?.[0] || "n/a", aEmo[aEmo.length - 1]?.[0] || "n/a", aEmo.length,
       aEmo.length ? formatValue(aEmo[aEmo.length - 1][1], "percent") : "n/a"],
      ["万得全A", wzqa[0]?.[0] || "n/a", wzqa[wzqa.length - 1]?.[0] || "n/a", wzqa.length,
       wzqa.length ? formatValue(wzqa[wzqa.length - 1][1], "index") : "n/a"]
    ];
    document.getElementById("emotion-table-a").innerHTML = tableHtml(["指标", "起点", "终点", "样本点数", "最新值"], rowsA);
  } else {
    document.getElementById("emotion-chart-a").innerHTML = "";
    document.getElementById("emotion-table-a").innerHTML = "";
  }

  // Chart 3: 港股情绪 vs 恒生指数 (dual axes)
  const hsi = pointsInRange("trend:恒生指数", start, end, { skipZero: true });
  if (hkEmo.length || hsi.length) {
    renderSvgLine("emotion-chart-hk", {
      title: "港股情绪与恒生指数",
      leftAxis: axisFromValues(hkEmo.map(p => p[1])),
      rightAxis: axisFromValues(hsi.map(p => p[1])),
      leftFormat: v => v.toFixed(1),
      rightFormat: v => formatValue(v, "index"),
      series: [
        { name: "港股情绪", points: hkEmo, color: "#C47B5A", axis: "left" },
        { name: "恒生指数", points: hsi, color: COLORS.green, axis: "right" }
      ]
    });
    const rowsHK = [
      ["港股情绪指数", hkEmo[0]?.[0] || "n/a", hkEmo[hkEmo.length - 1]?.[0] || "n/a", hkEmo.length,
       hkEmo.length ? formatValue(hkEmo[hkEmo.length - 1][1], "percent") : "n/a"],
      ["恒生指数", hsi[0]?.[0] || "n/a", hsi[hsi.length - 1]?.[0] || "n/a", hsi.length,
       hsi.length ? formatValue(hsi[hsi.length - 1][1], "index") : "n/a"]
    ];
    document.getElementById("emotion-table-hk").innerHTML = tableHtml(["指标", "起点", "终点", "样本点数", "最新值"], rowsHK);
  } else {
    document.getElementById("emotion-chart-hk").innerHTML = "";
    document.getElementById("emotion-table-hk").innerHTML = "";
  }
}

// ── Topics (Super Cycle) ────────────────────────────────────────────

const TOPIC_PERIOD = "FULL";

function topicTableHtml(tLabels, d1985, d2002, d2025) {
  const head = ["T+N", "1985", "2002", "2025"];
  const rows = tLabels.map((t, i) => [
    t,
    d1985[i] != null ? d1985[i].toFixed(2) : "—",
    d2002[i] != null ? d2002[i].toFixed(2) : "—",
    d2025[i] != null ? d2025[i].toFixed(2) : "—",
  ]);
  return tableHtml(head, rows);
}

function renderCategoricalLine(containerId, cfg) {
  // cfg: { tLabels: [...], series: [{name, color, lineWidth, data: [...]}], yMin, yMax, yLabel }
  const container = document.getElementById(containerId);
  if (!container) return;
  container.innerHTML = "";

  const { tLabels, series, yMin, yMax, yLabel } = cfg;
  const n = tLabels.length;
  if (n === 0) return;

  const M = { top: 20, right: 28, bottom: 44, left: 52 };
  const W = 920;
  const H = 400;
  const pw = W - M.left - M.right;
  const ph = H - M.top - M.bottom;
  const xScale = i => M.left + (i / (n - 1)) * pw;
  const yScale = v => M.top + ph * (1 - (v - yMin) / (yMax - yMin));

  let html = `<svg class="chart" viewBox="0 0 ${W} ${H}" xmlns="http://www.w3.org/2000/svg">`;

  // Y-axis grid + labels
  const ySteps = 5;
  for (let i = 0; i <= ySteps; i++) {
    const v = yMin + (yMax - yMin) * (i / ySteps);
    const y = yScale(v);
    html += `<line x1="${M.left}" x2="${W - M.right}" y1="${y}" y2="${y}" class="grid-line"/>`;
    html += `<text x="${M.left - 8}" y="${y + 4}" class="axis-text" text-anchor="end">${v}</text>`;
  }

  // Y-axis label
  html += `<text x="14" y="${M.top + ph/2}" class="axis-text" text-anchor="middle" transform="rotate(-90,14,${M.top + ph/2})">${yLabel || ""}</text>`;

  // X-axis labels
  const step = Math.max(1, Math.floor(n / 15));
  for (let i = 0; i < n; i++) {
    if (i % step !== 0 && i !== n - 1) continue;
    const x = xScale(i);
    html += `<text x="${x}" y="${H - 8}" class="axis-text" text-anchor="middle">${tLabels[i]}</text>`;
  }

  // Axis lines
  html += `<line x1="${M.left}" x2="${M.left}" y1="${M.top}" y2="${H - M.bottom}" class="axis-line"/>`;
  html += `<line x1="${M.left}" x2="${W - M.right}" y1="${H - M.bottom}" y2="${H - M.bottom}" class="axis-line"/>`;

  // Lines
  const activeSeries = series.filter(s => s.data.some(v => v != null));
  for (const s of activeSeries) {
    const pts = [];
    for (let i = 0; i < n; i++) {
      if (s.data[i] != null) pts.push([xScale(i), yScale(s.data[i])]);
    }
    if (pts.length < 2) continue;
    let d = "";
    for (let i = 0; i < pts.length; i++) {
      d += (i === 0 ? "M" : "L") + pts[i][0].toFixed(1) + "," + pts[i][1].toFixed(1);
    }
    html += `<path d="${d}" fill="none" stroke="${s.color}" stroke-width="${s.lineWidth || 2}" stroke-linejoin="round" stroke-linecap="round"/>`;
  }

  // Invisible hit areas for tooltip
  const bandW = pw / (n - 1);
  for (let i = 0; i < n; i++) {
    const cx = xScale(i);
    html += `<rect x="${cx - bandW/2}" y="${M.top}" width="${bandW}" height="${ph}" class="chart-hit" data-idx="${i}"/>`;
  }

  html += `</svg>`;
  container.innerHTML = html;

  // Tooltip
  let tooltip = null;
  container.querySelector("svg").addEventListener("mousemove", function(e) {
    const svg = this;
    const pt = svg.createSVGPoint();
    pt.x = e.clientX; pt.y = e.clientY;
    const sp = pt.matrixTransform(svg.getScreenCTM().inverse());
    const idx = Math.round((sp.x - M.left) / pw * (n - 1));
    const ci = Math.max(0, Math.min(n - 1, idx));

    if (tooltip) tooltip.remove();
    tooltip = document.createElementNS("http://www.w3.org/2000/svg", "g");

    const boxW = 140, boxH = 22 + activeSeries.length * 18;
    let bx = xScale(ci) + 12;
    let by = M.top + 8;
    if (bx + boxW > W - M.right) bx = xScale(ci) - boxW - 12;

    tooltip.innerHTML = `
      <rect x="${bx}" y="${by}" width="${boxW}" height="${boxH}" rx="5" class="tooltip-box"/>
      <text x="${bx + 10}" y="${by + 16}" class="tooltip-title">${tLabels[ci]}</text>
      ${activeSeries.map((s, si) => {
        const v = s.data[ci];
        return `<text x="${bx + 10}" y="${by + 34 + si * 18}" class="tooltip-text">${s.name}: ${v != null ? v.toFixed(2) : "—"}</text>`;
      }).join("")}
    `;
    svg.appendChild(tooltip);
  });
  container.querySelector("svg").addEventListener("mouseleave", function() {
    if (tooltip) { tooltip.remove(); tooltip = null; }
  });
}

function renderTopics() {
  const sc = DATA.super_cycle;
  if (!sc) {
    document.getElementById("topic-chart-dxy").innerHTML = `<div class="empty">专题数据文件未找到。请将 Super Dollar Scenario.xlsx 放置到 ~/Downloads/ 目录后重新生成看板。</div>`;
    return;
  }

  const { t_labels, dxy, dae } = sc;

  // Chart 1: DXY
  const dxyCfg = {
    tLabels: t_labels,
    yMin: 65, yMax: 100, yLabel: "Peak = 100",
    series: [
      { name: "DXY 1985", color: "#B8974A", lineWidth: 2, data: dxy["1985"] },
      { name: "DXY 2002", color: "#8B8581", lineWidth: 2, data: dxy["2002"] },
      { name: "DXY 2025", color: "#C44B3B", lineWidth: 2.8, data: dxy["2025"] },
    ]
  };
  renderCategoricalLine("topic-chart-dxy", dxyCfg);
  document.getElementById("topic-dxy-legend").innerHTML = dxyCfg.series.map(s =>
    `<span><span class="topic-legend-swatch" style="background:${s.color}"></span>${s.name}</span>`
  ).join("");
  document.getElementById("topic-table-dxy").innerHTML = topicTableHtml(t_labels, dxy["1985"], dxy["2002"], dxy["2025"]);

  // Chart 2: D/AE
  const daeCfg = {
    tLabels: t_labels,
    yMin: 65, yMax: 100, yLabel: "Peak = 100",
    series: [
      { name: "D/AE 1985", color: "#B8974A", lineWidth: 2, data: dae["1985"] },
      { name: "D/AE 2002", color: "#8B8581", lineWidth: 2, data: dae["2002"] },
      { name: "D/AE 2025", color: "#C44B3B", lineWidth: 2.8, data: dae["2025"] },
    ]
  };
  renderCategoricalLine("topic-chart-dae", daeCfg);
  document.getElementById("topic-dae-legend").innerHTML = daeCfg.series.map(s =>
    `<span><span class="topic-legend-swatch" style="background:${s.color}"></span>${s.name}</span>`
  ).join("");
  document.getElementById("topic-table-dae").innerHTML = topicTableHtml(t_labels, dae["1985"], dae["2002"], dae["2025"]);
}

function initTopics() {
  renderTopics();
}

// ── Navigation ──────────────────────────────────────────────────────

// ── FX Fixing (中间价) ──────────────────────────────────────────────

const FX_FIXING_SERIES = ["fx:usdcny-fixing", "fx:usdcny-spot", "fx:decomp-night-20d", "fx:decomp-day-20d"];
let fixingPeriod = "1Y";

function initFxFixing() {
  try {
  const container = document.getElementById("fixing-period");
  const options = [
    ["60D", "60天"], ["120D", "120天"], ["YTD", "YTD"],
    ["1Y", "1年"], ["3Y", "3年"], ["5Y", "5年"],
    ["10Y", "10年"], ["20150811", "Since 2015-08-11"],
  ];
  renderSegmented(container.id, options, fixingPeriod, val => {
    fixingPeriod = val;
    renderFxFixing();
  });
  renderFxFixing();
  } catch(e) {
    document.getElementById("fixing-chart-1").innerHTML = `<div class="empty">Init 错误: ${e.message || e}</div>`;
  }
}

function renderFxFixing() {
  try {
  const end = latestDate(FX_FIXING_SERIES);
  let start;
  if (fixingPeriod === "20150811") {
    start = "2015-08-11";
  } else if (fixingPeriod === "YTD") {
    const y = new Date(end).getFullYear();
    start = `${y}-01-01`;
  } else {
    start = startForPeriod(fixingPeriod, end);
  }

  const fixingRaw = pointsInRange("fx:usdcny-fixing", start, end);
  const spotRaw = pointsInRange("fx:usdcny-spot", start, end);
  // Only keep dates where both fixing and spot have data
  const spotMap = new Map(spotRaw.map(p => [p[0], p[1]]));
  const fixing = fixingRaw.filter(p => spotMap.has(p[0]));
  const spot = fixing.map(p => [p[0], spotMap.get(p[0])]);

  if (fixing.length === 0) {
    document.getElementById("fixing-chart-1").innerHTML = `<div class="empty">所选区间无数据</div>`;
    document.getElementById("fixing-table-1").innerHTML = "";
    document.getElementById("fixing-chart-2").innerHTML = "";
    document.getElementById("fixing-table-2").innerHTML = "";
    return;
  }

  // Chart 1: 中间价 vs 即期 + ±2% bands
  const band98 = fixing.map(p => [p[0], p[1] * 0.98]);
  const band102 = fixing.map(p => [p[0], p[1] * 1.02]);

  renderSvgLine("fixing-chart-1", {
    series: [
      { points: fixing, color: COLORS.blue, width: 2.2, axis: "left", name: "中间价" },
      { points: spot, color: COLORS.red, width: 1.8, axis: "left", name: "即期汇率" },
      { points: band98, color: COLORS.gold, width: 1.2, dash: "6,3", axis: "left", name: "中间价×98%", markers: false },
      { points: band102, color: COLORS.gold, width: 1.2, dash: "6,3", axis: "left", name: "中间价×102%", markers: false },
    ],
    leftAxis: axisFromValues([...fixing.map(p=>p[1]), ...spot.map(p=>p[1]), ...band98.map(p=>p[1]), ...band102.map(p=>p[1])]),
    leftFormat: v => v.toFixed(3),
  });

  document.getElementById("fixing-table-1").innerHTML = tableHtml(
    ["日期", "中间价", "即期汇率", "偏离(%)", "±2%下轨", "±2%上轨"],
    fixing.map((p, i) => {
      const dev = spot[i] ? ((spot[i][1] / p[1] - 1) * 100) : null;
      return [
        p[0], p[1].toFixed(4), spot[i] ? spot[i][1].toFixed(4) : "—",
        dev != null ? dev.toFixed(2) + "%" : "—",
        (p[1] * 0.98).toFixed(4), (p[1] * 1.02).toFixed(4),
      ];
    }).slice(-1).reverse()
  );

  // Chart 2: 20-day rolling decomposition
  const night20 = pointsInRange("fx:decomp-night-20d", start, end);
  const day20 = pointsInRange("fx:decomp-day-20d", start, end);

  if (night20.length > 0 || day20.length > 0) {
    renderSvgLine("fixing-chart-2", {
      series: [
        { points: night20, color: COLORS.blue, width: 2, axis: "left", name: "夜盘中间价调整(20MA)" },
        { points: day20, color: "#C47B5A", width: 2, axis: "left", name: "日盘交易变动(20MA)" },
      ],
      leftAxis: axisFromValues([...night20.map(p=>p[1]), ...day20.map(p=>p[1])]),
      leftFormat: v => v.toFixed(0) + " pips",
    });

    document.getElementById("fixing-table-2").innerHTML = tableHtml(
      ["日期", "夜盘调整(pips)", "日盘变动(pips)", "合计(pips)"],
      night20.map((p, i) => {
        const day = day20[i] ? day20[i][1] : null;
        const sum = p[1] != null && day != null ? p[1] + day : null;
        return [p[0], p[1].toFixed(0), day != null ? day.toFixed(0) : "—", sum != null ? sum.toFixed(0) : "—"];
      }).slice(-1).reverse()
    );
  }
  } catch(e) {
    document.getElementById("fixing-chart-1").innerHTML = `<div class="empty">渲染错误: ${e.message}</div>`;
    console.error("renderFxFixing error:", e);
  }
}

// ── FX Cost (套保成本) ──────────────────────────────────────────────

const FX_COST_SERIES = [
  "fx:cny-hedge-1m","fx:cny-hedge-3m","fx:cny-hedge-6m","fx:cny-hedge-1y",
  "fx:cnh-hedge-1m","fx:cnh-hedge-3m","fx:cnh-hedge-6m","fx:cnh-hedge-1y",
  "fx:cny-hedge-ann-1m","fx:cny-hedge-ann-3m","fx:cny-hedge-ann-6m","fx:cny-hedge-ann-1y",
  "fx:cnh-hedge-ann-1m","fx:cnh-hedge-ann-3m","fx:cnh-hedge-ann-6m","fx:cnh-hedge-ann-1y",
  "fx:cny-bond-1y","fx:usd-bond-1y",
];

let costContract = "both";
let costModes = new Set(["latest"]);
let cost3mPeriod = "1Y";
let cost1yPeriod = "1Y";
let cost3mShowCny = true;
let cost3mShowCnh = true;
let cost3mSmoothing = "raw";
let cost1yShowCny = true;
let cost1yShowCnh = true;
let cost1ySmoothing = "raw";

const TENORS = ["1M","3M","6M","1Y"];
const COST_MODE_LABELS = { latest: "最新值", "5d": "5日均值", ytd: "YTD均值", soy: "年初水平", range: "指定区间" };

function costAggregate(seriesId, mode, endDate, rangeStart, rangeEnd) {
  const end = endDate || latestDate(FX_COST_SERIES);
  let start, endLimit = end;
  if (mode === "latest") {
    const r = nearestOnOrBefore(seriesId, end);
    return r ? r[1] : null;
  } else if (mode === "5d") {
    start = toISO(addDays(toDate(end), -5));
  } else if (mode === "ytd") {
    start = `${new Date(end).getFullYear()}-01-01`;
  } else if (mode === "soy") {
    start = `${new Date(end).getFullYear()}-01-01`;
    endLimit = `${new Date(end).getFullYear()}-01-31`;
  } else if (mode === "range" && rangeStart && rangeEnd) {
    start = rangeStart; endLimit = rangeEnd;
  }
  if (start) {
    const pts = pointsInRange(seriesId, start, endLimit);
    if (pts.length > 0) {
      return pts.reduce((s, p) => s + p[1], 0) / pts.length;
    }
  }
  return null;
}

const MODE_COLORS = { latest: COLORS.blue, "5d": "#C47B5A", ytd: COLORS.green, soy: COLORS.purple, range: COLORS.gold };

function rolling5dMA(points) {
  const result = [];
  for (let i = 0; i < points.length; i++) {
    const start = Math.max(0, i - 4);
    let sum = 0;
    for (let j = start; j <= i; j++) sum += points[j][1];
    result.push([points[i][0], sum / (i - start + 1)]);
  }
  return result;
}

function renderTermStructure(containerId, modesData, contract) {
  // modesData: [{key, label, tenors: [{label, cny, cnh, cnyAnn, cnhAnn}]}]
  const container = document.getElementById(containerId);
  if (!container) return;
  container.innerHTML = "";

  const tenors = modesData[0]?.tenors || [];
  if (!tenors.length) { container.innerHTML = `<div class="empty">无数据</div>`; return; }

  // Collect all annualized values for Y range
  const allVals = [];
  for (const md of modesData) {
    for (const d of md.tenors) {
      if (contract !== "cnh" && d.cnyAnn != null) allVals.push(d.cnyAnn);
      if (contract !== "cny" && d.cnhAnn != null) allVals.push(d.cnhAnn);
    }
  }
  if (allVals.length === 0) { container.innerHTML = `<div class="empty">无数据</div>`; return; }

  // Compute legend entry count for margin
  let legendEntries = 0;
  for (const md of modesData) {
    if (contract !== "cnh") legendEntries++;
    if (contract !== "cny") legendEntries++;
  }
  const legendRows = Math.ceil(legendEntries / 3);
  const M = { top: 12 + legendRows * 22, right: 28, bottom: 36, left: 58 };
  const W = 640, H = M.top + 260 + M.bottom;
  const pw = W - M.left - M.right;
  const ph = H - M.top - M.bottom;

  const yMin = Math.floor(Math.min(...allVals) * 10000) / 10000;
  const yMax = Math.ceil(Math.max(...allVals) * 10000) / 10000;
  const yPad = Math.max(0.001, (yMax - yMin) * 0.2);
  const yLo = yMin - yPad, yHi = yMax + yPad;
  const yScale = v => M.top + ph * (1 - (v - yLo) / (yHi - yLo));
  const xPos = i => M.left + (i / (tenors.length - 1)) * pw;

  let html = `<svg class="chart" viewBox="0 0 ${W} ${H}" style="min-width:${W}px;height:${H}px" xmlns="http://www.w3.org/2000/svg">`;

  // Grid + Y labels
  const ySteps = 4;
  for (let i = 0; i <= ySteps; i++) {
    const v = yLo + (yHi - yLo) * (i / ySteps);
    const y = yScale(v);
    html += `<line x1="${M.left}" x2="${W - M.right}" y1="${y}" y2="${y}" class="grid-line"/>`;
    html += `<text x="${M.left - 8}" y="${y + 4}" class="axis-text" text-anchor="end">${(v * 100).toFixed(2)}%</text>`;
  }
  // X labels
  for (let i = 0; i < tenors.length; i++) {
    html += `<text x="${xPos(i)}" y="${H - 6}" class="axis-text" text-anchor="middle">${tenors[i].label}</text>`;
  }
  // Axis lines
  html += `<line x1="${M.left}" x2="${M.left}" y1="${M.top}" y2="${H - M.bottom}" class="axis-line"/>`;
  html += `<line x1="${M.left}" x2="${W - M.right}" y1="${H - M.bottom}" y2="${H - M.bottom}" class="axis-line"/>`;

  // Lines + dots + data labels per mode (年化套保成本 only)
  // CNY = solid line, CNH = dashed line (same mode color)
  const labelOffsets = {};
  for (let mi = 0; mi < modesData.length; mi++) {
    const md = modesData[mi];
    const color = MODE_COLORS[md.key] || COLORS.muted;

    const groups = [];
    if (contract !== "cnh") groups.push({ key: "cnyAnn", label: "CNY", dash: false });
    if (contract !== "cny") groups.push({ key: "cnhAnn", label: "CNH", dash: true });

    for (const g of groups) {
      let d = "";
      for (let i = 0; i < tenors.length; i++) {
        if (md.tenors[i][g.key] == null) continue;
        const x = xPos(i), y = yScale(md.tenors[i][g.key]);
        d += (d === "" ? "M" : "L") + x.toFixed(1) + "," + y.toFixed(1);
      }
      if (d) {
        const dashAttr = g.dash ? ' stroke-dasharray="6,4"' : '';
        html += `<path d="${d}" fill="none" stroke="${color}" stroke-width="2.2" stroke-linejoin="round"${dashAttr}/>`;
        for (let i = 0; i < tenors.length; i++) {
          if (md.tenors[i][g.key] == null) continue;
          const x = xPos(i), y = yScale(md.tenors[i][g.key]);
          html += `<circle cx="${x.toFixed(1)}" cy="${y.toFixed(1)}" r="3.5" fill="${color}" stroke="#fff" stroke-width="1"/>`;
          const valStr = (md.tenors[i][g.key] * 100).toFixed(2) + "%";
          const lblKey = g.key + i;
          if (!(lblKey in labelOffsets)) labelOffsets[lblKey] = 0;
          const stack = labelOffsets[lblKey]++;
          const offY = -14 - stack * 16;
          html += `<text x="${x.toFixed(1)}" y="${(y + offY).toFixed(1)}" class="data-label" text-anchor="middle" fill="${color}">${valStr}</text>`;
        }
      }
    }
  }

  // Legend — one entry per mode × contract (solid=CNY, dashed=CNH)
  let legendIdx = 0;
  for (let mi = 0; mi < modesData.length; mi++) {
    const md = modesData[mi];
    const color = MODE_COLORS[md.key] || COLORS.muted;
    const entries = [];
    if (contract !== "cnh") entries.push({ label: md.label + " CNY", dash: false });
    if (contract !== "cny") entries.push({ label: md.label + " CNH", dash: true });
    for (const e of entries) {
      const row = Math.floor(legendIdx / 3), col = legendIdx % 3;
      const lx = M.left + col * 195, ly = 16 + row * 22;
      const dashAttr = e.dash ? ' stroke-dasharray="6,4"' : '';
      html += `<line x1="${lx}" y1="${ly}" x2="${lx + 18}" y2="${ly}" stroke="${color}" stroke-width="2.2"${dashAttr}/>`;
      html += `<text x="${lx + 24}" y="${ly + 4}" class="legend-text">${e.label}</text>`;
      legendIdx++;
    }
  }

  html += `</svg>`;
  container.innerHTML = html;
}

function initFxCost() {
  const contractOptions = [
    ["both", "CNY + CNH"], ["cny", "仅 CNY"], ["cnh", "仅 CNH"],
  ];
  const periodOptions = [
    ["60D", "60天"], ["120D", "120天"], ["YTD", "YTD"],
    ["1Y", "1年"], ["3Y", "3年"], ["5Y", "5年"],
    ["10Y", "10年"], ["20150811", "Since 2015-08-11"],
  ];

  renderSegmented("cost-contract", contractOptions, costContract, val => { costContract = val; renderFxCost(); });

  // Checkbox-based multi-mode selection
  const modeContainer = document.getElementById("cost-mode");
  modeContainer.innerHTML = "";
  for (const [key, label] of [["latest","最新值"],["5d","5日均值"],["ytd","YTD均值"],["soy","年初水平"],["range","指定区间"]]) {
    const lbl = document.createElement("label");
    lbl.className = "mode-check";
    lbl.innerHTML = `<input type="checkbox" value="${key}" ${costModes.has(key) ? "checked" : ""}><span>${label}</span>`;
    lbl.querySelector("input").addEventListener("change", ev => {
      if (ev.target.checked) costModes.add(key); else costModes.delete(key);
      if (costModes.size === 0) { ev.target.checked = true; costModes.add(key); return; }
      document.getElementById("cost-custom-wrap").style.display = costModes.has("range") ? "block" : "none";
      renderFxCost();
    });
    modeContainer.appendChild(lbl);
  }
  document.getElementById("cost-custom-wrap").style.display = costModes.has("range") ? "block" : "none";

  renderSegmented("cost-3m-period", periodOptions, cost3mPeriod, val => { cost3mPeriod = val; renderFxCost(); });
  renderSegmented("cost-1y-period", periodOptions, cost1yPeriod, val => { cost1yPeriod = val; renderFxCost(); });

  // 3M contract checkboxes
  (function() {
    const container = document.getElementById("cost-3m-contract");
    container.innerHTML = "";
    for (const [key, label, ref] of [["cny","CNY",()=>cost3mShowCny],["cnh","CNH",()=>cost3mShowCnh]]) {
      const lbl = document.createElement("label");
      lbl.className = "mode-check";
      const checked = key === "cny" ? cost3mShowCny : cost3mShowCnh;
      lbl.innerHTML = `<input type="checkbox" value="${key}" ${checked ? "checked" : ""}><span>${label}</span>`;
      lbl.querySelector("input").addEventListener("change", ev => {
        if (key === "cny") cost3mShowCny = ev.target.checked; else cost3mShowCnh = ev.target.checked;
        if (!cost3mShowCny && !cost3mShowCnh) {
          if (key === "cny") { cost3mShowCny = true; ev.target.checked = true; }
          else { cost3mShowCnh = true; ev.target.checked = true; }
          return;
        }
        renderFxCost();
      });
      container.appendChild(lbl);
    }
  })();

  // 3M smoothing toggle
  renderSegmented("cost-3m-ma", [["raw","当日值"],["5dma","5日均值"]], cost3mSmoothing, val => {
    cost3mSmoothing = val; renderFxCost();
  });

  // 1Y contract checkboxes
  (function() {
    const container = document.getElementById("cost-1y-contract");
    container.innerHTML = "";
    for (const [key, label] of [["cny","CNY"],["cnh","CNH"]]) {
      const lbl = document.createElement("label");
      lbl.className = "mode-check";
      const checked = key === "cny" ? cost1yShowCny : cost1yShowCnh;
      lbl.innerHTML = `<input type="checkbox" value="${key}" ${checked ? "checked" : ""}><span>${label}</span>`;
      lbl.querySelector("input").addEventListener("change", ev => {
        if (key === "cny") cost1yShowCny = ev.target.checked; else cost1yShowCnh = ev.target.checked;
        if (!cost1yShowCny && !cost1yShowCnh) {
          if (key === "cny") { cost1yShowCny = true; ev.target.checked = true; }
          else { cost1yShowCnh = true; ev.target.checked = true; }
          return;
        }
        renderFxCost();
      });
      container.appendChild(lbl);
    }
  })();

  // 1Y smoothing toggle
  renderSegmented("cost-1y-ma", [["raw","当日值"],["5dma","5日均值"]], cost1ySmoothing, val => {
    cost1ySmoothing = val; renderFxCost();
  });

  ["cost-range-start","cost-range-end"].forEach(id => {
    document.getElementById(id).addEventListener("change", () => renderFxCost());
  });

  renderFxCost();
}

function renderFxCost() {
  try {
  const end = latestDate(FX_COST_SERIES);
  const rangeS = document.getElementById("cost-range-start")?.value || "";
  const rangeE = document.getElementById("cost-range-end")?.value || "";

  // ── Chart 1: Multi-mode Term Structure ──
  const activeModes = [...costModes].filter(m => m !== "range" || (rangeS && rangeE));
  const modesData = [];
  for (const mode of activeModes) {
    const tenors = [];
    for (const tenor of TENORS) {
      const tl = tenor.toLowerCase();
      const cnySid = `fx:cny-hedge-${tl}`;
      const cnhSid = `fx:cnh-hedge-${tl}`;
      const cnyAnnSid = `fx:cny-hedge-ann-${tl}`;
      const cnhAnnSid = `fx:cnh-hedge-ann-${tl}`;
      tenors.push({
        label: tenor,
        cny: costAggregate(cnySid, mode, end, rangeS, rangeE),
        cnh: costAggregate(cnhSid, mode, end, rangeS, rangeE),
        cnyAnn: costAggregate(cnyAnnSid, mode, end, rangeS, rangeE),
        cnhAnn: costAggregate(cnhAnnSid, mode, end, rangeS, rangeE),
      });
    }
    modesData.push({ key: mode, label: COST_MODE_LABELS[mode] || mode, tenors });
  }
  renderTermStructure("cost-chart-1", modesData, costContract);

  // Table for term structure — one row per mode+tenor combo
  const tsHeader = ["取值方式","期限","CNY年化套保","CNH年化套保"];
  const tsRows = [];
  for (const md of modesData) {
    for (const d of md.tenors) {
      tsRows.push([
        md.label, d.label,
        d.cnyAnn != null ? (d.cnyAnn*100).toFixed(2)+"%" : "—",
        d.cnhAnn != null ? (d.cnhAnn*100).toFixed(2)+"%" : "—",
      ]);
    }
  }
  document.getElementById("cost-table-1").innerHTML = tableHtml(tsHeader, tsRows);

  // ── Chart 2: 3M Annualized Time Series ──
  const s3m = cost3mPeriod === "20150811" ? "2015-08-11"
    : cost3mPeriod === "YTD" ? `${new Date(end).getFullYear()}-01-01`
    : startForPeriod(cost3mPeriod, end);
  let cny3mAnn = pointsInRange("fx:cny-hedge-ann-3m", s3m, end);
  let cnh3mAnn = pointsInRange("fx:cnh-hedge-ann-3m", s3m, end);
  if (cost3mSmoothing === "5dma") {
    cny3mAnn = rolling5dMA(cny3mAnn);
    cnh3mAnn = rolling5dMA(cnh3mAnn);
  }

  if (cny3mAnn.length > 0 || cnh3mAnn.length > 0) {
    const s3mSeries = [];
    if (cost3mShowCny) {
      s3mSeries.push({ points: cny3mAnn, color: COLORS.blue, width: 2, axis: "left", name: "CNY 3M套保成本(年化)" });
    }
    if (cost3mShowCnh) {
      s3mSeries.push({ points: cnh3mAnn, color: "#C47B5A", width: 2, axis: "left", name: "CNH 3M套保成本(年化)" });
    }
    if (s3mSeries.length > 0) {
      const all3mVals = s3mSeries.flatMap(s => s.points.map(p => p[1]));
      renderSvgLine("cost-chart-2", {
        series: s3mSeries,
        leftAxis: axisFromValues(all3mVals),
        leftFormat: v => (v*100).toFixed(2) + "%",
      });
    }
    document.getElementById("cost-table-2").innerHTML = "";
  }

  // ── Chart 3: 1Y Annualized Hedge vs CN-US 1Y Spread ──
  const s1y = cost1yPeriod === "20150811" ? "2015-08-11"
    : cost1yPeriod === "YTD" ? `${new Date(end).getFullYear()}-01-01`
    : startForPeriod(cost1yPeriod, end);
  let cny1yAnn = pointsInRange("fx:cny-hedge-ann-1y", s1y, end);
  let cnh1yAnn = pointsInRange("fx:cnh-hedge-ann-1y", s1y, end);
  if (cost1ySmoothing === "5dma") {
    cny1yAnn = rolling5dMA(cny1yAnn);
    cnh1yAnn = rolling5dMA(cnh1yAnn);
  }
  const cnBond1y = pointsInRange("fx:cny-bond-1y", s1y, end);
  const usBond1y = pointsInRange("fx:usd-bond-1y", s1y, end);

  const cnMap = new Map(cnBond1y.map(p => [p[0], p[1]]));
  const usMap = new Map(usBond1y.map(p => [p[0], p[1]]));
  let spread = [];
  for (const [d, cv] of cnBond1y) {
    const uv = usMap.get(d);
    if (uv != null) spread.push([d, cv - uv]);
  }
  if (cost1ySmoothing === "5dma") {
    spread = rolling5dMA(spread);
  }

  if (cny1yAnn.length > 0 || cnh1yAnn.length > 0) {
    const s1ySeries = [{ points: spread, color: COLORS.green, width: 1.8, axis: "right", name: "中美1Y利差" }];
    if (cost1yShowCny) {
      s1ySeries.unshift({ points: cny1yAnn, color: COLORS.blue, width: 2, axis: "left", name: "CNY 1Y套保成本(年化)" });
    }
    if (cost1yShowCnh) {
      s1ySeries.unshift({ points: cnh1yAnn, color: "#C47B5A", width: 2, axis: "left", name: "CNH 1Y套保成本(年化)" });
    }
    const leftVals = s1ySeries.filter(s => s.axis === "left").flatMap(s => s.points.map(p => p[1]));
    if (leftVals.length > 0) {
      renderSvgLine("cost-chart-3", {
        series: s1ySeries,
        leftAxis: axisFromValues(leftVals),
        leftFormat: v => (v*100).toFixed(2) + "%",
        rightAxis: axisFromValues(spread.map(p => p[1])),
        rightFormat: v => v.toFixed(2) + "pp",
      });
    }
    document.getElementById("cost-table-3").innerHTML = "";
  }
  } catch(e) {
    document.getElementById("cost-chart-1").innerHTML = `<div class="empty">渲染错误: ${e.message}</div>`;
    console.error("renderFxCost error:", e);
  }
}

// ── Navigation: hierarchical primary → sub-tab ──────────────────────

let _currentGroup = null;

const TAB_GROUPS = {
  trend:  { label: "走势看板", children: [] },
  review: { label: "涨跌复盘", children: ["returns","rates","fx"] },
  equity: { label: "权益看板", children: ["returns","valuation","emotion"] },
  fxwatch:{ label: "外汇看板", children: ["fx","spread","fixing","cost","topics"] },
};

const VIEW_LABELS = {
  trend: "走势看板", returns: "股票涨跌", rates: "利率涨跌",
  fx: "汇率涨跌", spread: "中美利差", valuation: "估值看板",
  emotion: "市场情绪", topics: "专题图表", fixing: "中间价", cost: "套保成本",
};

function viewGroup(viewName) {
  // Find which group(s) a view belongs to
  const groups = [];
  for (const [gk, gv] of Object.entries(TAB_GROUPS)) {
    if (gk === viewName || gv.children.includes(viewName)) groups.push(gk);
  }
  return groups.length > 0 ? groups[0] : null;
}

function switchToView(viewName) {
  document.querySelectorAll(".view").forEach(v => v.classList.remove("active"));
  const view = document.getElementById(`view-${viewName}`);
  if (view) view.classList.add("active");

  // Resolve group: prefer current group if it contains this view
  let group = _currentGroup;
  if (!group || !TAB_GROUPS[group]) {
    group = viewGroup(viewName);
  } else {
    const cfg = TAB_GROUPS[group];
    const inGroup = cfg.children.length === 0 ? group === viewName : cfg.children.includes(viewName);
    if (!inGroup) group = viewGroup(viewName);
  }

  // Highlight primary tab
  document.querySelectorAll(".primary-tabs .tab").forEach(b => b.classList.remove("active"));
  if (group) {
    const pt = document.querySelector(`.primary-tabs .tab[data-group="${group}"]`);
    if (pt) pt.classList.add("active");
  }

  // Show & highlight sub-tabs
  const subBar = document.getElementById("sub-tabs");
  const cfg = group ? TAB_GROUPS[group] : null;
  if (cfg && cfg.children.length > 0) {
    subBar.innerHTML = cfg.children.map(v =>
      `<button class="tab" data-view="${v}">${VIEW_LABELS[v] || v}</button>`
    ).join("");
    subBar.classList.add("visible");
    document.querySelectorAll("#sub-tabs .tab").forEach(b => b.classList.remove("active"));
    const st = document.querySelector(`#sub-tabs .tab[data-view="${viewName}"]`);
    if (st) st.classList.add("active");
  } else {
    subBar.classList.remove("visible");
    subBar.innerHTML = "";
  }
}

function switchToGroup(groupName) {
  const cfg = TAB_GROUPS[groupName];
  if (!cfg) return;
  _currentGroup = groupName;
  const defaultView = cfg.children.length > 0 ? cfg.children[0] : groupName;
  switchToView(defaultView);
}

function initTabs() {
  // Primary tabs
  document.querySelectorAll(".primary-tabs .tab").forEach(btn => {
    btn.addEventListener("click", () => switchToGroup(btn.dataset.group));
  });
  // Sub-tabs (delegated — re-rendered dynamically)
  document.getElementById("sub-tabs").addEventListener("click", e => {
    const btn = e.target.closest(".tab");
    if (btn && btn.dataset.view) switchToView(btn.dataset.view);
  });
  // Cover cards — direct view navigation
  document.querySelectorAll(".cover-card").forEach(card => {
    card.addEventListener("click", () => switchToView(card.dataset.view));
  });
}

function init() {
  document.getElementById("freshness").textContent = `生成时间 ${DATA.generated_at}`;
  document.getElementById("cover-date").textContent = DATA.generated_at;
  document.getElementById("sources").innerHTML = `数据源：本地 SQLite 数据库。绘图与计算口径：历史取数为 0 的记录视为缺失值，并用上一期有效读数填补；开头连续为 0 且没有上一期有效读数的记录跳过。`;
  initTabs();
  initTrend();
  initReturns();
  initRates();
  initFx();
  initSpread();
  initValuation();
  initEmotion();
  initTopics();
  initFxFixing();
  initFxCost();
}
init();
  </script>
</body>
</html>
"""


def render_dashboard(db_path, output_path):
    payload = build_payload(db_path)
    html = HTML_TEMPLATE.replace("__DATA__", json.dumps(payload, ensure_ascii=False, separators=(",", ":")))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding="utf-8")
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Generate the interactive Martin Morning Brief dashboard.")
    parser.add_argument("--db", default=str(DEFAULT_DB))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = parser.parse_args()
    output = render_dashboard(Path(args.db), Path(args.output))
    print(f"Generated {output}")


if __name__ == "__main__":
    main()
